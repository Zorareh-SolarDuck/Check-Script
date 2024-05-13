# -*- coding: utf-8 -*-
"""
Created on Thu Jul 29 14:39:18 2021

@author: Sander
"""
import OrcFxAPI as ofx
import numpy as np
import numpy.matlib
import math
from scipy import optimize
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils  import get_column_letter
import matplotlib.pyplot as plt
import copy
import re
import pandas as pd
import os
import warnings

##Classes in which input is defined 

## Some classes are formed based on the parametric input data from the excel
## The classes can be used to adjust all the input for the model

class MassObject:
    def __init__(self, Mass, CoG, InertiaTensor=np.zeros([3,3])):
        self.Mass=Mass
        self.CoG=CoG
        if np.shape(InertiaTensor)==(3,):
            self.I=np.zeros([3,3])
            self.I[0,0]=InertiaTensor[0]
            self.I[1,1]=InertiaTensor[1]
            self.I[2,2]=InertiaTensor[2]
            # self.I
        else:
            self.I=np.array(InertiaTensor) ## This tensor is always around the CoG
        
    def set_filled_triangle_interia(self, EdgeLength):
        Length=EdgeLength*np.sqrt(3)/2
        self.I[0,0]= (self.Mass)/18*Length**2
        self.I[1,1]= (self.Mass)/18*Length**2
        self.I[2,2]= 1/36*(self.Mass)*(3*EdgeLength**2)
        
    def set_triangle_border_interia(self, EdgeLengthOut, EdgeLengthIn):
        rho=self.Mass/(EdgeLengthOut**2*3**0.5/2-EdgeLengthIn**2*3**0.5/2)
        m1=EdgeLengthOut**2*3**0.5/2*rho
        m0=EdgeLengthIn**2*3**0.5/2*rho
        self.I[0,0]= m1/18*(EdgeLengthOut*(3)**0.5/2)**2-m0/18*(EdgeLengthIn*(3)**0.5/2)**2
        self.I[1,1]=self.I[0,0]
        self.I[2,2]= 1/36*m1*(3*EdgeLengthOut**2)-1/36*m0*(3*EdgeLengthIn**2)
    
    def set_cylinder_inertia(self, Diameter, Height):
        self.I[0,0]=1/12*self.Mass*(6*(Diameter/2)**2+Height**2)
        self.I[1,1]=1/12*self.Mass*(6*(Diameter/2)**2+Height**2)
        self.I[2,2]=self.Mass*(Diameter/2)**2

class Platform:
    def __init__(self):
        self.EdgeLength=25#+2.5
        self.EdgeLengthOuter=30
        self.MassT1=20.1
        self.MassT2=23.3
        self.TrussBottomZ=0.223  
        self.Height=3.5+self.TrussBottomZ
                      
        self.Colour=12632256    
        self.Colour_L4=10485760
        self.TrussLineName = ['TrussInT1','TrussInT2','TrussOutT1','TrussOutT2']
        self.Split=False
        self.PendulumOn=False
        self.AlternatingCoupling=False
        self.AjustMassOfOuterPlatforms=True
        self.AdjustMassOfEdgePlatforms=False
        self.BoxTruss = False
        self.WindShielding = False
        self.PlaceSkid = False
        self.CoeReduceRatio = 1
        self.TrussWidth = 1.8
        self.TrussHeight = 3.4
        self.TussFloaterIntersection = 8.921
         
        self.BoxTrussEIx = 2.18e6 # kNm2
        self.BoxTrussEIy = 7.77e6 # kNm2
        self.BoxTrussEA = 2.69e6 # kN
        self.BoxTrussGJ = 0.56E6 # kNm
        self.TrussProjectedAreaX = 31.4 # m2
        self.TrussProjectedAreaY = 13.2 # m2
        
     

    def calc_details_basic(self):
        self.Length=self.EdgeLength*np.sqrt(3)/2
        self.GeoCoG=[self.Length/3, 0, self.Height/2] 
        
        #### This function should be recreated for when the platform is not split to create 1 rigid body mass and properties
            
        # self.Truss.set_triangle_border_interia(self.EdgeLength, 23)
        # self.PanelsT1.set_triangle_border_interia(self.EdgeLength+2,self.EdgeLength)
        # self.PanelsT2.set_triangle_border_interia(self.EdgeLength+2,self.EdgeLength)
        
        ## Total platform T1 is calcualated
        # self.T0=self.add_mass_objects([self.PanelsAndGrating, self.Truss])
        
        self.AirGapProbesX=np.array([0,0,0, self.Length/2, self.Length/2, self.Length/3,self.Length])
        self.AirGapProbesY=np.array([0,self.EdgeLength/2, -self.EdgeLength/2, self.EdgeLength/4, -self.EdgeLength/4, 0,0])       
        self.AirGapProbesZ=np.zeros(self.AirGapProbesX.size)+self.TrussBottomZ #### Updated for new results
            
        self.DrawingVertexX=[0,0,0,0,self.Length,self.Length]
        self.DrawingVertexY=[self.EdgeLength/2, self.EdgeLength/2, -self.EdgeLength/2, -self.EdgeLength/2, 0,0]
        self.DrawingVertexZ=[self.Height,0,self.Height,0,self.Height,0]      

        Length=self.EdgeLengthOuter*3**0.5/2
        dX=np.tan(30*np.pi/180)*(self.EdgeLengthOuter-self.EdgeLength)/2         
        
        self.DrawingVertexX_L4=np.array([0,0,0,0,Length,Length])-dX
        self.DrawingVertexY_L4=[self.EdgeLengthOuter/2, self.EdgeLengthOuter/2, -self.EdgeLengthOuter/2, -self.EdgeLengthOuter/2, 0,0]
        self.DrawingVertexZ_L4=[self.Height+0.5,self.Height,self.Height+0.5,self.Height,self.Height+0.5,self.Height]    
        
    
    def add_airgap_probes_xy_outer(self,EdgeLength):
        Length=EdgeLength*3**0.5/2
        dX=np.tan(30*np.pi/180)*(EdgeLength-self.EdgeLength)/2
        self.AirGapProbesX=np.append(self.AirGapProbesX,np.array([0,0,0, Length/2, Length/2,  Length])-dX)
        self.AirGapProbesY=np.append(self.AirGapProbesY,np.array([0,EdgeLength/2, -EdgeLength/2, EdgeLength/4, -EdgeLength/4, 0]))
        self.AirGapProbesZ=np.array(np.append(self.AirGapProbesZ,self.AirGapProbesZ+self.Height-self.TrussBottomZ))
        
    def add_mass_objects(self,ListOfObjects):
        Out=MassObject(0.0,np.zeros(3))
        Mass=[]
        CoGs=[]
        Is=[]
        for L in ListOfObjects:
            Mass.append(L.Mass)
            CoGs.append(L.CoG)
            Is.append(L.I)
        Masses=np.array(Mass)
        CoGs=np.array(CoGs)
        Is=np.array(Is)
               
        Out.Mass=np.sum(Masses)
        Out.CoG=np.sum(Mass*np.transpose(CoGs), axis=1)/np.sum(Masses)
        Rs=CoGs-Out.CoG
        I_PATs=Masses*np.array([[Rs[:,1]**2+Rs[:,2]**2, -Rs[:,0]*Rs[:,1]     ,   -Rs[:,0]*Rs[:,2]],
                       [-Rs[:,1]*Rs[:,0]     , Rs[:,0]**2+Rs[:,2]**2,   -Rs[:,1]*Rs[:,2]],
                       [-Rs[:,2]*Rs[:,0]     , -Rs[:,2]*Rs[:,1]     , Rs[:,0]**2+Rs[:,1]**2]])
    
        I_PAT=np.sum(I_PATs, axis=2)
        Out.I=I_PAT+np.sum(Is,axis=0)
        return Out
        
    def find_alpha_cog(self, Alpha):
        return np.array([np.tan(30*np.pi/180)*Alpha, self.EdgeLength/2-Alpha, self.ZTopFloater])    
    
    def calc_beam_conlocations(self, BeamAlpha):
        BeamConLocA=[0,0,[np.sin(60*np.pi/180)*BeamAlpha, self.EdgeLength/2-np.cos(60*np.pi/180)*BeamAlpha, self.BeamHeight]]
        BeamConLocB=[[0,self.EdgeLength/2-BeamAlpha,self.BeamHeight],0,0]
        
        if self.BoxTruss:
            # BeamConLocA=[[self.TrussWidth/2,-(self.EdgeLength/2-BeamAlpha), self.BeamHeight]]
            # BeamConLocB=[[self.TrussWidth/2,self.EdgeLength/2-BeamAlpha,self.BeamHeight]]
            BeamConLocA=[[self.TrussWidth/2,-self.TussFloaterIntersection, self.BeamHeight]]
            BeamConLocB=[[self.TrussWidth/2,self.TussFloaterIntersection,self.BeamHeight]]
            
            
        # if self.BoxTruss:
        #     self.BeamEquivLen = self.EdgeLength - 2*self.TrussWidth/2/np.tan(np.deg2rad(30))
        #     BeamEquiAlpha = BeamAlpha - self.TrussWidth/2/np.tan(np.deg2rad(60))
        #     BeamConLocA=[[self.TrussWidth/2,-(self.BeamEquivLen/2-BeamEquiAlpha), self.BeamHeight]]
        #     BeamConLocB=[[self.TrussWidth/2,self.BeamEquivLen/2-BeamEquiAlpha,self.BeamHeight]]
            # BeamConLocA=[0,0,[np.sin(60*np.pi/180)*BeamAlpha,self.EdgeLength/2-BeamAlpha/2, self.BeamHeight]]
            # BeamConLocB=[[0,self.EdgeLength/2-BeamAlpha,self.BeamHeight],0,0]
        return BeamConLocA, BeamConLocB
    
    def calc_drawing_vertex_split(self, CornerLength):
        DrawingVertexX0=np.array([0,0,0,0,np.sin(60*np.pi/180)*CornerLength,np.sin(60*np.pi/180)*CornerLength])
        DrawingVertexY0=self.EdgeLength/2-np.array([0,0, CornerLength, CornerLength,np.cos(60*np.pi/180)*CornerLength,np.cos(60*np.pi/180)*CornerLength])
        return DrawingVertexX0,DrawingVertexY0
                         
    def split_platform_in_three_initiate(self):
        self.BeamAlpha_T1=3
        self.BeamAlpha_T2=3
        
        # self.TranslationT2=np.array([1.2,-1.2/np.tan(30*np.pi/180),0])
        self.TranslationT2=np.zeros(4)
        
        # self.CornerMassT1=4.185/3
        # self.CornerMassT2=3.6/3
        
    def split_platform_in_three_calc_details(self): 
        
        self.TrussAndPanels=self.add_mass_objects([self.TrussLayer1,self.TrussLayer2, self.PanelsAndGrating])
        # self.BeamHeight=self.TrussAndPanels.CoG[2]
        
        self.CoGAlpha_T1=self.BeamAlpha_T1/2
        self.CoGAlpha_T2=self.BeamAlpha_T2/2
        
        self.CoG0_T1=self.find_alpha_cog(self.CoGAlpha_T1)
        self.CoG0_T2=self.find_alpha_cog(self.CoGAlpha_T2)+self.TranslationT2## Triangle is translated to new position
        
        # #### The 0 Corners are used as basis to create a new values that is used in each loop in the construction phase
        # self.T1Corner0=MassObject(self.MassCornersTotal/3,self.CoG0_T1)
        # self.T1Corner0.set_filled_triangle_interia(self.CoGAlpha_T1*2)
        # self.T1Corner=MassObject(self.MassCornersTotal/3,self.CoG0_T1)
        # self.T1Corner.set_filled_triangle_interia(self.CoGAlpha_T1*2)
        
        # self.T2Corner0=MassObject(self.MassCornersTotal/3, self.CoG0_T2)
        # self.T2Corner0.set_filled_triangle_interia(self.CoGAlpha_T2*2)
        # self.T2Corner=MassObject(self.MassCornersTotal/3, self.CoG0_T2)
        # self.T2Corner.set_filled_triangle_interia(self.CoGAlpha_T2*2)
                
        # self.AirGapProbesX=np.array([0,self.Length/2,0])
        # self.AirGapProbesY=np.array([0,0,self.EdgeLength/2]) ## These need to be rotated with the body
        if not self.BoxTruss:
            self.BeamHeight=self.TrussAndPanels.CoG[2]
            self.NodesXYZ_T1=[[0,0,self.BeamHeight]]
            self.NodesXYZ_T2=[[self.TranslationT2[0],0,self.BeamHeight]]
            self.BeamConnectionsA=['Node_0_0_0', 'Node_0_0_0','Platform_0_0']
            self.BeamConnectionsB=['Platform_0_0', 'Node_0_240_0', 'Node_0_240_0']
            
            self.BeamConLocA_T1, self.BeamConLocB_T1=self.calc_beam_conlocations(self.BeamAlpha_T1)
            self.BeamConLocA_T2, self.BeamConLocB_T2=self.calc_beam_conlocations(self.BeamAlpha_T2)
            self.BeamConLocA_T2[2]=self.BeamConLocA_T2[2]+self.TranslationT2
            self.BeamConLocB_T2[0]=self.BeamConLocB_T2[0]+self.TranslationT2
            
            self.BeamTypes=['TrussOutT','TrussInT','TrussOutT']
            self.BeamLengths_T1=[self.EdgeLength/2-self.BeamAlpha_T1,self.EdgeLength/2, self.EdgeLength/2-self.BeamAlpha_T1]
            self.BeamLengths_T2=[self.EdgeLength/2-self.BeamAlpha_T2+self.TranslationT2[1],(self.EdgeLength+self.TranslationT2[1]*2)/2,self.EdgeLength/2-self.BeamAlpha_T2+self.TranslationT2[1]]
            self.BeamAzimuth=[90,30,-30]
            self.BeamSegments=2
            self.DrawingVertexZ0=np.array([self.Height,0,self.Height,0,self.Height,0])
            
            self.DrawingVertexX0_T1, self.DrawingVertexY0_T1=self.calc_drawing_vertex_split(self.BeamAlpha_T1)
            self.DrawingVertexX0_T2, self.DrawingVertexY0_T2=self.calc_drawing_vertex_split(self.BeamAlpha_T2)
            self.DrawingVertexX0_T2=self.DrawingVertexX0_T2+self.TranslationT2[0]
            self.DrawingVertexY0_T2=self.DrawingVertexY0_T2+self.TranslationT2[1]
            
            self.Split=True  
            self.set_truss_mass()
        else:
            self.BeamHeight=1.7
            self.BeamConnectionsA=['Platform_0_120']
            self.BeamConnectionsB=['Platform_0_0']
            self.BeamConLocA_T1, self.BeamConLocB_T1=self.calc_beam_conlocations(self.BeamAlpha_T1)
            self.BeamConLocA_T2, self.BeamConLocB_T2=self.calc_beam_conlocations(self.BeamAlpha_T2)
            self.BeamConLocA_T2[0]=self.BeamConLocA_T2[0]+self.TranslationT2
            self.BeamConLocB_T2[0]=self.BeamConLocB_T2[0]+self.TranslationT2
            self.BeamTypes=['BoxTrussT']
            # self.BeamLengths_T1=[self.EdgeLength-2*self.BeamAlpha_T1]
            # self.BeamLengths_T2=[self.EdgeLength-2*self.BeamAlpha_T2+self.TranslationT2[1]*2]
            
            self.BeamLengths_T1=[self.TussFloaterIntersection*2]
            self.BeamLengths_T2=[self.TussFloaterIntersection*2]
            
            self.BeamAzimuth=[90]
            self.BeamSegments=2
            self.DrawingVertexZ0=np.array([self.Height,0,self.Height,0,self.Height,0])
            
            self.DrawingVertexX0_T1, self.DrawingVertexY0_T1=self.calc_drawing_vertex_split(self.BeamAlpha_T1)
            self.DrawingVertexX0_T2, self.DrawingVertexY0_T2=self.calc_drawing_vertex_split(self.BeamAlpha_T2)
            self.DrawingVertexX0_T2=self.DrawingVertexX0_T2+self.TranslationT2[0]
            self.DrawingVertexY0_T2=self.DrawingVertexY0_T2+self.TranslationT2[1]
            
            self.Split=True  
            self.set_truss_mass()
        # if self.AdjustMassOfEdgePlatforms:
        #     self.set_truss_mass_edge()
               
    def set_truss_mass(self):
        AreaIn=self.EdgeLength**2*3**0.5/2
        AreaTotal=self.EdgeLengthOuter**2*3**0.5/2
        PanelsOutMass=self.MassPanelsAndGrating
        PanelsInMass=0
        
        #### Hardcoded for indentical platforms
        if not self.BoxTruss:
            self.TrussLineName = self.TrussLineName # ['TrussInT1','TrussInT2','TrussOutT1','TrussOutT2']
            # self.TrussLineName = ['TrussInT1','TrussInT2','TrussOutT1','TrussOutT2']
            self.OFBeamMassL1=self.TrussLayer1.Mass+PanelsOutMass+PanelsInMass/2
            self.OFBeamMassL2=self.TrussLayer2.Mass+PanelsInMass/2
            self.LineMassPerLength={}
            self.LineMassPerLength[self.TrussLineName[0]]=self.OFBeamMassL2/(np.sum(self.BeamLengths_T1[1])*3)
            self.LineMassPerLength[self.TrussLineName[1]]=self.OFBeamMassL2/(np.sum(self.BeamLengths_T2[1])*3)
            self.LineMassPerLength[self.TrussLineName[2]]=self.OFBeamMassL1/(np.sum([self.BeamLengths_T1[0],self.BeamLengths_T1[2]])*3)
            self.LineMassPerLength[self.TrussLineName[3]]=self.OFBeamMassL1/(np.sum([self.BeamLengths_T2[0],self.BeamLengths_T2[2]])*3)
            
            self.BeamCoGZOffset={}
            self.BeamCoGZOffset[self.TrussLineName[0]]=0
            self.BeamCoGZOffset[self.TrussLineName[1]]=0
            self.BeamCoGZOffset[self.TrussLineName[2]]=0
            self.BeamCoGZOffset[self.TrussLineName[3]]=0
            
            ### set truss properties for skid platform
            self.LineMassPerLengthSkid={}
            self.TrussLineNameSkid = ['TrussInT2Skid','TrussOutT2Skid']
            self.OFBeamMassL1Skid=self.TrussLayer1.Mass 
            self.OFBeamMassL2Skid=self.TrussLayer2.Mass
            self.LineMassPerLengthSkid[self.TrussLineNameSkid[0]]=self.OFBeamMassL2Skid/(np.sum(self.BeamLengths_T2[1])*3)
            self.LineMassPerLengthSkid[self.TrussLineNameSkid[1]]=self.OFBeamMassL1Skid/(np.sum([self.BeamLengths_T2[0],self.BeamLengths_T2[2]])*3)

        else:
            self.LineMassPerLength={}
            self.BeamCoGZOffset={}
            self.TrussLineName = ['BoxTrussT1','BoxTrussT2']
            self.OFBeamMassL1=self.TrussLayer1.Mass+PanelsOutMass+PanelsInMass/2
            self.LineMassPerLength[self.TrussLineName[0]]=self.OFBeamMassL1/(np.sum([self.BeamLengths_T1[0]])*3)
            self.LineMassPerLength[self.TrussLineName[1]]=self.OFBeamMassL1/(np.sum([self.BeamLengths_T1[0]])*3)
            self.BeamCoGZOffset[self.TrussLineName[0]]=1.485
            self.BeamCoGZOffset[self.TrussLineName[1]]=1.365
        
            ### set truss properties for skid platform
            self.LineMassPerLengthSkid={}
            self.TrussLineNameSkid = ['TrussOutT2Skid']
            self.OFBeamMassL1Skid=self.TrussLayer1.Mass 
            self.OFBeamMassL2Skid=self.TrussLayer2.Mass
            # self.LineMassPerLengthSkid[self.TrussLineNameSkid[0]]=self.OFBeamMassL2Skid/(np.sum(self.BeamLengths_T2[1])*3)
            self.LineMassPerLengthSkid[self.TrussLineNameSkid[0]]=self.OFBeamMassL1Skid/(np.sum(self.BeamLengths_T2[0])*3)
            
            self.TrussNormalDragLiftDiameter = self.TrussProjectedAreaX / (self.TussFloaterIntersection*2)
            self.TrussAxialDragLiftDiameter = self.TrussProjectedAreaY / (self.TussFloaterIntersection*2)
        
        #### set truss properties for wind shielding
        # self.TrussLineNameWindShielding = ['TrussInT1WindShielding','TrussInT2WindShielding','TrussOutT1WindShielding','TrussOutT2WindShielding']


    def set_coupling_inertia_contribution(self, Coupling, Compose, CouplingZPos=False):
        ## Extra Panel masses have been added to the coupling as they will have a similar mass distribution
        ## This could be improved at a later stage
        ## Calc mean coupling positions
        Position=[]
        for k in Coupling.Position.keys():
            Position.append(Coupling.Position[k])
        Position.append(Coupling.Position['Center'])
        Position.append(Coupling.Position['Center'])
        Position=np.mean(np.array(Position),axis=0)
        Position[0]=Coupling.Position['Mid'][0]-Coupling.Gap/2
        
        if CouplingZPos:
            Position[2]=CouplingZPos    
            
        self.CouplingCoG=Position
        
        if self.Split:
            self.T1Corner={}
            self.T1CornerOut={}
            self.T2Corner={}
            self.T2CornerOut={}
            self.TopFloater={}


            for j in [0,120,240]:
                self.TopFloater[str(j)]=MassObject(self.MassTopFloaterTotal/3, Compose.translate_coordinates(self, self.CoG0_T1, j))
                self.TopFloater[str(j)].set_cylinder_inertia(2.25,self.Height)

                
                self.T1Corner[str(j)]=MassObject(self.MassCornersTotal/3,Compose.translate_coordinates(self, self.CoG0_T1, j))
                self.T1Corner[str(j)].set_filled_triangle_interia(self.CoGAlpha_T1*2)
                Coupling1_T1=MassObject(self.MassCouplingsTotal_T1/6,Compose.translate_coordinates(self, Position, j))
                Coupling2_T1=MassObject(self.MassCouplingsTotal_T1/6, Compose.translate_coordinates(self, Position*np.array([1,-1,1]), 240+j))
                self.T1CornerOut[str(j)]=self.add_mass_objects([self.T1Corner[str(j)],Coupling1_T1, Coupling2_T1,self.TopFloater[str(j)]])

                
                self.T2Corner[str(j)]=MassObject(self.MassCornersTotal/3,Compose.translate_coordinates(self, self.CoG0_T1, j))
                self.T2Corner[str(j)].set_filled_triangle_interia(self.CoGAlpha_T2*2)
                Coupling1_T2=MassObject(self.MassCouplingsTotal_T2/6,Compose.translate_coordinates(self, Position, j))
                Coupling2_T2=MassObject(self.MassCouplingsTotal_T2/6, Compose.translate_coordinates(self, Position*np.array([1,-1,1]), 240+j))
                self.T2CornerOut[str(j)]=self.add_mass_objects([self.T2Corner[str(j)],Coupling1_T2, Coupling2_T2,self.TopFloater[str(j)]])

        #### Section below was used to calc the coupling mass contribution to a rigid platform 
        #### Section is not neccesary anymore as more accurate esimates can be made from FEM models
        #### If it needs to be used a summation of the section above would be more useful
            
        # self.CouplingTotal=MassObject(self.MassCouplingTotal,np.zeros(3))        
        # ## Calc coupling positions        
        # Positions=[]
        # for i in [1,-1]:
        #     for idir in [0,120,240]:
        #         Pos=Compose.translate_coordinates(self, Position*[1,i,1], idir)
        #         Positions.append(Pos)   
        # self.CouplingTotal.CoG=np.mean(Positions,axis=0)
        # Dxyz=Positions-self.CouplingTotal.CoG
        # Ixyz=np.matmul(self.CouplingTotal.Mass/6*Dxyz**2, [[0,1,1],
        #                                 [1,0,1],
        #                                 [1,1,0]])
        # Ixyz=np.sum(Ixyz, axis=0)
        # self.CouplingTotal.I[0,0]=Ixyz[0]
        # self.CouplingTotal.I[1,1]=Ixyz[1]
        # self.CouplingTotal.I[2,2]=Ixyz[2]
        
        if self.Split:
            if self.BoxTruss:
                self.OFBeamMassL2 = 0
            self.MassNoFloaters_T1=self.OFBeamMassL1+self.OFBeamMassL2+self.T1CornerOut['0'].Mass*3+self.MassNodesTotal
            self.MassNoFloaters_T2=self.OFBeamMassL1+self.OFBeamMassL2+self.T2CornerOut['0'].Mass*3+self.MassNodesTotal    
            

            
    def find_platforms(self,Layout):
        NoneZeroRows = []
        NonZeroRowIndex = []
        NoneZeroPos = []
        PlatformOrientations = []
        EdgePlatforms = []
        AllPlatforms = []
        EdgePlatformsTypes = []
        EdgePlatformsOrientation = []
        CornerPlatforms = []
        CornerPlatformsTypes = []
        CornerPlatformOrientations = []
        ####Step1 find the none zero rows and columns in Layout.Matrix
        # Matrix = Layout.Matrix[np.any(Layout.Matrix,axis=1)][:,np.any(Layout.Matrix,axis=0)]
        for i, row in enumerate(Layout.Matrix):
            if any(row):
                NonZeroRowIndex.append(i)
                NoneZeroRows.append(row)
        
        ####Step2 find the positions of each none zero elemnt, positions are positions in the matrix
        for ir, nrow in enumerate(NoneZeroRows):
            NoneZeroPos.append([j + Layout.Matrix.shape[1]*NonZeroRowIndex[ir] for j, val in enumerate(nrow) if val != 0])
            # NoneZeroPos.append([j + Mat.shape[1]*ir for j, val in enumerate(nrow) if val != 0])
        #### Step3 find the all/edge elements
        #### for edge paltform 
        #### only in the first and the last rows all the elements are accounted
        #### in the rest rows only the first two and the last two elements are accounted   
        for ip, pfn in enumerate(NoneZeroPos):
            # print(ip)
            NoneZeroElements = list(filter(lambda x: x != 0, NoneZeroRows[ip]))
            PlatformOrientations.append([0 if ori==1 else 180 for ori in NoneZeroElements])
            AllPlatforms.append(pfn)
            if ip == 0 or ip == (len(NoneZeroPos)-1): #### first and last row
                EdgePlatforms.append(pfn)
                EdgePlatformsTypes.append([rj for rj in NoneZeroElements])
                EdgePlatformsOrientation.append(PlatformOrientations[ip])
                CornerPlatforms.append([jj for iip, jj in enumerate(pfn) if iip in [0,1, len(pfn)-2, len(pfn)-1]])
                CornerPlatformsTypes.append([NoneZeroElements[iip] for iip, jj in enumerate(pfn) if iip in [0,1, len(pfn)-2, len(pfn)-1]])
                CornerPlatformOrientations.append([jj for iip, jj in enumerate(PlatformOrientations[ip]) if iip in [0,1, len(pfn)-2, len(pfn)-1]])
                
            else:
                # EdgePlatforms.append([jj for iip, jj in enumerate(pfn) if iip == 0 or iip == 1 or iip == (len(pfn)-1) or iip == (len(pfn)-2)])
                # EdgePlatformsTypes.append([rj for ir, rj in enumerate(NoneZeroElements) if ir == 0 or ir == 1 or ir == (len(NoneZeroElements)-1) or ir == (len(NoneZeroElements)-2) ])
                EdgePlatforms.append([jj for iip, jj in enumerate(pfn) if iip in [0,1,len(pfn)-2,len(pfn)-1]])
                EdgePlatformsTypes.append([rj for ir, rj in enumerate(NoneZeroElements) if ir in [0,1,len(NoneZeroElements)-2,len(NoneZeroElements)-1]])
                EdgePlatformsOrientation.append([rj for ir, rj in enumerate(PlatformOrientations[ip]) if ir in [0,1,len(PlatformOrientations[ip])-2,len(PlatformOrientations[ip])-1]])
                if ip in [NoneZeroPos.index(max(NoneZeroPos, key=len)),NoneZeroPos.index(max(NoneZeroPos, key=len))+1]:
                    CornerPlatforms.append([jj for iip, jj in enumerate(pfn) if iip in [0,len(pfn)-1]])
                    CornerPlatformsTypes.append([NoneZeroElements[iip] for iip, jj in enumerate(pfn) if iip in [0, len(pfn)-1]])
                    CornerPlatformOrientations.append([jj for iip, jj in enumerate(PlatformOrientations[ip]) if iip in [0, len(pfn)-1]])
        self.AllPlatforms = [item for sublist in AllPlatforms for item in sublist]
        self.EdgePlatforms = [item for sublist in EdgePlatforms for item in sublist]
        # self.Edges = EdgePlatforms#[EdgePlatforms[0][1:-1],EdgePlatforms[0][0]+EdgePlatforms[1][0:2]+EdgePlatforms[2][0:2],]
    
        self.InnerPlatforms = [item for item in self.AllPlatforms if item not in self.EdgePlatforms]
        self.EdgePlatformsTypes = [item for sublist in EdgePlatformsTypes for item in sublist]
        self.NoneZeroRows = NoneZeroRows
        
        self.InCorners = CornerPlatforms
        self.CornerPlatforms = [item for sublist in CornerPlatforms for item in sublist]
        self.CornerPlatformsArray = CornerPlatforms
        self.CornerPlatformsTypes = [item for sublist in CornerPlatformsTypes for item in sublist]
        self.EdgePlatformsT1 = [self.EdgePlatforms[ielem] for ielem, elem in enumerate(self.EdgePlatformsTypes) if elem == 1]
        self.EdgePlatformsT2 = [self.EdgePlatforms[ielem] for ielem, elem in enumerate(self.EdgePlatformsTypes) if elem == 2]
        
        self.Orientations = [item for sublist in PlatformOrientations for item in sublist]
        self.EdgePlatformsOrientation = [item for sublist in EdgePlatformsOrientation for item in sublist]
        self.CornerPlatformOrientations = [item for sublist in CornerPlatformOrientations for item in sublist]
        
        
class Pendulum:
        def __init__(self, Platform):
            self.PosCenter=Platform.CoG*[1,1,0]+[0,0,-4.5]   
            self.Mass={}
            self.Mass['Center']=0.1
            self.Mass['Pend']=1.5
            self.Density={}
            self.Density['Center']=7 ## t/m^3 steel plate
            self.Density['Pend']=3 ## t/m^3 concrete mass
            self.StiffnessRope=10000 ## EA random estimate ~ half the stiffness of the mooring 
            self.LenPedulum=5 ## this can be tuned  
            self.Cd=0.8
            self.Ca=0.8
            
        def calc_details(self, Platform, Floater):
            self.LenLink=np.linalg.norm([np.linalg.norm(Platform.CoG[0:2]- Floater.PosBottom[0:2]),self.PosCenter[2]])
            self.Volume={}
            self.Height={}
            self.DragArea={}
            for i in self.Mass.keys():
                self.Volume[i]=self.Mass[i]/self.Density[i]
                self.Height[i]=(self.Volume[i])**(1/3)
                self.DragArea[i]=(self.Volume[i])**(2/3)
            self.Pos={}
            self.Pos['Center']=self.PosCenter
            self.Pos['Pend']=self.PosCenter-[0,0, self.LenPedulum]
        
class Floater: 
        def __init__(self):
            self.Diameter=1.83#*(14.65/16.8)**(2/3)
            self.Height=6.5
            self.CdAxial=0.6
            self.CdNormalAboveWL=0.7 ## WL stands for waterline
            self.CdNormalUnderWL=0.8 ## WL stand for waterline
            self.CdNormal=0.8 ## This can be used if no difference between above and below waterline is considered

            self.CaAxial=0.27
            self.CaNormal=0.8
            self.WaveDampAxial=0.4 #kNs/m ## This is for the whole floater
            self.WaveDampNormal=1 #kNs/m
            self.WaveDampNormalSegment=self.WaveDampNormal/2
            self.PosBottom=np.array([0.915,6.45,-7.5])
            self.NumCylinders=4
            self.Mass=4.293/3
            self.NumberOfFloaters=3
            self.Colour=32768
            self.EndShape='RoundCap'
            self.ZConnectionPoint=0 #### give the connection point of the floater, in some cases this need to be update to match FEM anaylses
            self.VBottomCone=1.489
            # self.CoG = np.array([0,0,self.Height/2])
            
        def calc_details(self, PlatformMass):
            self.CoG=np.array([0,0,self.Height/2])
            self.AreaAxial=np.pi*((self.Diameter/2)**2)
            self.Draft=(PlatformMass+self.Mass*3)/1.025/(self.AreaAxial*self.NumberOfFloaters)
            self.MassOb=MassObject(self.Mass, self.CoG)
            self.MassOb.set_cylinder_inertia(self.Diameter, self.Height)
            self.DisplacedMassPlatform=PlatformMass+self.Mass*3
            
            # self.I=np.zeros([3,3])
            # self.I[0,0]=1/12*self.Mass*(6*(self.Diameter/2)**2+self.Height**2)
            # self.I[1,1]=1/12*self.Mass*(6*(self.Diameter/2)**2+self.Height**2)
            # self.I[2,2]=self.Mass*(self.Diameter/2)**2
            
        def set_drag_coefficients_straight_floater(self, TopSections, BottomSections):
            self.NumCylinders=TopSections+BottomSections
            self.CylinderLength=np.array([(self.Height-self.Draft)/TopSections for i in range(TopSections)]+ [self.Draft/BottomSections for i in range(BottomSections)])
            self.CylinderDiameter=np.array([self.Diameter for i in range(self.NumCylinders)])
            self.CdNormal=np.array([self.CdNormalAboveWL for i in range(TopSections)]+[self.CdNormalUnderWL for i in range(BottomSections)])
        
        def set_axial_Ca_based_on_flat_plate(self):
            CaDNVC205=0.64
            ReferenceVolume=4/3*(self.Diameter/2)**3*np.pi
            self.AxialAddedMass=ReferenceVolume*CaDNVC205/2*1.025
            self.CaAxial=self.AxialAddedMass/(self.Draft*self.AreaAxial*1.025)
            
        # def scale(self, ScaleOrigin, ScaleNew):
        #     R=ScaleNew/ScaleOrigin
        #     self.Diameter=self.Diameter*R
        #     self.Height=self.Height*R
        #     self.WaveDampAxial=self.WaveDampAxial*R**3
        #     self.WaveDampNormal=self.WaveDampNormal*R**3
        #     self.PosBottom=self.PosBottom*R
        #     self.Mass=self.Mass*R**3
        #     self.Ixx=self.Ixx*R**5
        #     self.Iyy=self.Iyy*R**5
        #     self.Izz=self.Izz*R**5
        #     self.Draft=self.Draft*R
        #     self.CoG=np.array(self.CoG)*R
            
        def apply_marine_growth(self, DensityGrowth=1.325, Thickness=0.1, Cd=1.05):
            DiaOrig=self.Diameter
            ## Assume that floater draft has bee set to satisfy the mass
            DispOrig=self.AreaAxial*self.Draft*1.025 #t
            self.Diameter=self.Diameter+2*Thickness
            self.DenistyGrowth=DensityGrowth 
            self.ThicknessGrowth=Thickness
            ## Cd axial still needs to be adjusted accordingly
            AreaOrig=np.pi*DiaOrig**2/4
            Area=np.pi*self.Diameter**2/4
            dA=Area-AreaOrig
            self.CdNormalUnderWL=Cd
            # self.Draft=(Area*Thickness-dA*(DensityGrowth/1.025-1)+DispOrig/1.025)/(Area-dA*(DensityGrowth/1.025))+Thickness
            self.Draft=(Area*Thickness*DensityGrowth/1.025+DispOrig/1.025)/(Area-dA*(DensityGrowth/1.025))
            self.Height=self.Height+Thickness
            self.AreaAxial=Area
            self.MassGrowth=DensityGrowth*(dA*self.Draft+Thickness*AreaOrig)
            self.CoG=[0,0,((self.Draft/2)*self.MassGrowth+self.CoG[2]*self.Mass)/(self.MassGrowth+self.Mass)]
            self.Mass=self.Mass+self.MassGrowth
            self.PosBottom=self.PosBottom+[0,0,-Thickness]
            self.set_axial_Ca_based_on_flat_plate()
            self.set_drag_coefficients()
            self.CylinderDiameter[0:2]=DiaOrig

        # def apply_taper(self, TopDiameter, BottomDiameter, HeightMainSection, Segments):
        #     ## Top and bottom diameters are for the main floater section            
        #     ## Note the difference beteen Cd air and Cd water is hardcoded here!! And wave damping is hardcoded
        #     self.WaveDampNormalSegment=self.WaveDampNormal/3            
        #     self.NumCylinders=Segments+1 #### +1 for the cone section
        #     # BottomDiameter=TopDiameter-self.Height*np.tan(TaperAngle*np.pi/180)*2
        #     CylinderHeight=(HeightMainSection/Segments)
        #     ## Quick estimate of cylinder diameters it is not copletely accurate
        #     CylinderHeights=np.arange(10)
        #     self.CylinderDiameter=np.interp(CylinderHeights, [0,9],[TopDiameter, BottomDiameter])            
        #     self.CdNormal=np.ones(Segments)*self.CdNormalAboveWL
        #     self.CdNormal[-3:]=self.CdNormalUnderWL
        #     print('CdNormalHardcoded')
        #     self.CylinderLength=np.array([CylinderHeight for i in range(Segments)])
        #     self.AreaAxial=1/4*np.pi*BottomDiameter**2
        #     self.TopDiameter=TopDiameter
        #     self.BottomDiameter=BottomDiameter
        #     #### Bottom cone dimensions are mentioned here this is hardcoded watch out 
        #     self.VBottomCone=1/3*np.pi*0.409*(0.8**2+0.8*0.3+0.3**2)
        #     EqAreaBottomCone=self.VBottomCone/0.409
        #     EqDiameterBottomCone=(4*EqAreaBottomCone/np.pi)**0.5
        #     self.CylinderDiameter=np.append(self.CylinderDiameter,EqDiameterBottomCone)
        #     self.CylinderLength=np.append(self.CylinderLength,0.409)
        #     self.CdNormal=np.append(self.CdNormal,self.CdNormalUnderWL)
        #     self.Height=np.sum(self.CylinderLength)
            
        def apply_double_taper(self, TopDiameter, MidDiameter, BottomDiameter, HeightTopSection, HeightMidSection, SegmentsTop, SegmentsMid):
            ## Top and bottom diameters are for the main floater section            
            ## Note the difference beteen Cd air and Cd water is hardcoded here!! And wave damping is hardcoded
            self.WaveDampNormalSegment=self.WaveDampNormal/4            
            self.MidDiameter=MidDiameter
            self.BottomDiameter=BottomDiameter
            self.TopDiameter=TopDiameter
            self.HeightTopSection=HeightTopSection           
            #### Bottom cone dimensions are mentioned here this is hardcoded watch out 
            self.HeightBottomCone=self.HeightBottomCone
            if self.EndShape=='RoundCap':
                # self.VBottomCone=1.489
                self.VBottomCone=self.VBottomCone 
                self.DiameterBottomCone=self.calc_diameter_cylinder(self.VBottomCone,self.HeightBottomCone)
            elif Floater.EndShape=='Cone':
                self.DiameterBottomCone=0.6
                self.VBottomCone=self.calc_partial_cone_volume(BottomDiameter,self.DiameterBottomCone, self.HeightBottomCone)
            
            # self.Height=0.409+HeightMidSection
            self.HeightTopSection=HeightTopSection
            
        def calc_partial_cone_volume(self,D1,D2, h):
            return 1/3*np.pi*h*((D1/2)**2+D1/2*D2/2+(D2/2)**2)
        
        def calc_volume_cylinder(self, D,h):
            return np.pi/4*D**2*h
        
        def calc_diameter_cylinder(self, V, h):
            return (4*V/(np.pi*h))**0.5
        
        def get_cone_properties(self,DTop, DBottom, Height, Segments):
            Htot=np.linspace(0,Height, Segments+1)
            Hseg=Height/Segments
            r=DBottom/2
            Diameters=[]
            Vtot=[]
            Vsegement=[]
            EqDiameters=[]
            for ih, h in enumerate(Htot):
                R=h*(DTop/2-DBottom/2)/Height+r
                Diameters.append(R*2)
                Vtot.append(self.calc_partial_cone_volume(Diameters[-1], Diameters[0], h))
                if ih>0:
                    Vsegement.append(self.calc_partial_cone_volume(Diameters[-1], Diameters[-2], Hseg))
                    EqDiameters.append(self.calc_diameter_cylinder(Vsegement[-1],Hseg))
            CylinderLengths=np.ones(Segments)*Hseg
            OutVol=[Htot, Diameters, Vtot]
            OutOF=[np.flip(EqDiameters), CylinderLengths, Vsegement]
            
            return OutVol, OutOF 
        
      
        def calc_taper_draft_and_inertia(self, PlatformMass, Platform):
            ## This can be run after applying the taper to the floater
            ## This determines draft using middle section only
            self.DisplacedMassPlatform=PlatformMass+self.Mass*3
            
            OutVol, OutOF = self.get_cone_properties(self.MidDiameter, self.BottomDiameter, self.Height-self.HeightBottomCone, 19)
            
            self.VFloaterdH=np.array(OutVol[2])+self.VBottomCone
            self.DiameterdH=np.array(OutVol[1])
            
            # plt.plot(np.array(Vtot)+self.VBottomCone,np.array(htot)+0.409)
            # plt.xlabel('Volume[m^3]')
            # plt.ylabel('Draft [m]')
            # plt.title('Single Floater Displacement Plot')
            # plt.minorticks_on()
            # plt.grid(which='major',color='k')
            if self.HeightTopSection:
                self.Draft=np.interp(self.DisplacedMassPlatform/3/1.025-self.VBottomCone,OutVol[2],OutVol[0])+self.HeightBottomCone
                self.MassObTop=MassObject(self.MassTop, [0,0,self.HeightTopSection/2+self.Height])
                self.MassObTop.set_cylinder_inertia((self.TopDiameter+self.MidDiameter)/2, self.HeightTopSection)
                self.MassObBottom=MassObject(self.MassBottom, [0,0,self.Height/2])
                self.MassObBottom.set_cylinder_inertia((self.MidDiameter+self.BottomDiameter)/2, self.Height)
                #### The tempex in the floater has been added here                
                self.MassObTempex=MassObject(self.MassTempex,[0,0,self.ZTempex])
                self.MassObTempex.set_cylinder_inertia(self.MidDiameter, self.HeightTempex)
                self.MassOb=Platform.add_mass_objects([self.MassObTop, self.MassObBottom, self.MassObTempex])
                
                
        def apply_taper_marine_growth(self, Platform, DensityGrowth=1.325, Thickness=0.1, Cd=1.05):
            #### This function must be run after the apply double taper function without changing floater heights 
            #### This function must be run also for 0 marine growth 
            t=Thickness
            if self.EndShape=='Cone':
                self.VBottomConeMG=self.calc_partial_cone_volume(self.BottomDiameter+2*t, self.DiameterBottomCone+2*t, self.HeightBottomCone)+ self.calc_volume_cylinder(self.DiameterBottomCone+2*t,2*t)
            elif self.EndShape=='RoundCap':
                self.VBottomConeMG=self.calc_volume_cylinder(self.DiameterBottomCone+2*t,t+self.HeightBottomCone)
                
                
            self.MidDiameterMG=self.MidDiameter+2*t
            self.BottomDiameterMG=self.BottomDiameter+2*t
                        
            OutVol, OutOF = self.get_cone_properties(self.MidDiameterMG, self.BottomDiameterMG, self.Height-self.HeightBottomCone, 19)
                      
            self.VFloaterdHMG=np.array(OutVol[2])+self.VBottomConeMG
            self.MassdHMG=(self.VFloaterdHMG-self.VFloaterdH)*DensityGrowth
            
            MassBalance=self.VFloaterdHMG*1.025-self.MassdHMG-self.DisplacedMassPlatform/3
            Draft=np.interp(0,MassBalance, OutVol[0])+self.HeightBottomCone
            MassMG=np.interp(0,MassBalance,self.MassdHMG)
            # plt.plot(np.array(OutVol[0])+0.40,MassBalance)
            
            DiameterMG=np.interp(0,MassBalance,OutVol[1])
            DiameterNMG=np.interp(0,MassBalance,self.DiameterdH)
            
            OutLengths=[]
            OutDiameters=[]
            
            a, OutOF = self.get_cone_properties(self.TopDiameter,self.MidDiameter, self.HeightTopSection, 3)
            OutLengths.extend(OutOF[1])
            OutDiameters.extend(OutOF[0])
            a, OutOF = self.get_cone_properties(self.MidDiameter,DiameterNMG, self.Height-Draft, 3)
            OutLengths.extend(OutOF[1])
            OutDiameters.extend(OutOF[0])
            a, OutOF = self.get_cone_properties(DiameterMG, self.BottomDiameterMG, Draft-self.HeightBottomCone, 3)
            OutLengths.extend(OutOF[1])
            OutDiameters.extend(OutOF[0])
            #### The last addon has been changed to account for a round cap instead of just the cone
            if self.EndShape=='Cone':
                a, OutOF = self.get_cone_properties(self.BottomDiameterMG, self.DiameterBottomCone+2*t , self.HeightBottomCone+t, 1)
            elif self.EndShape=='RoundCap':
                OutOF[1]=[self.HeightBottomCone]
                OutOF[0]=[self.DiameterBottomCone]
                
            OutLengths.extend(OutOF[1])
            OutDiameters.extend(OutOF[0])
    
            self.CylinderDiameter=OutDiameters
            self.CylinderLength=OutLengths
            
            # self.Height=np.sum(Lengths[3:])
            self.Draft=Draft+t
            self.AreaAxial=np.pi/4*(self.BottomDiameter+2*t)**2
            #### define different drag coe in inner paltform floaters
            self.CdNormalTopFloaterIn = self.CdNormalAboveWL*Platform.CoeReduceRatio #### smaller drag coe inner platform
            # self.CdNormalIn=np.append(np.ones(6)*self.CdNormalTopFloaterIn, np.ones(4)*Cd)
            self.CdNormalIn=np.concatenate([np.ones(3)*self.CdNormalTopFloaterIn,np.ones(3)*self.CdNormalAboveWL, np.ones(4)*Cd])
            #### 
            self.CdNormal=np.append(np.ones(6)*self.CdNormalAboveWL, np.ones(4)*Cd)
            self.NumCylinders=len(self.CylinderDiameter)
                        
            self.MassObMG=MassObject(MassMG, np.array([0,0,self.Draft/2]))
            self.MassObMG.set_cylinder_inertia(self.Diameter+2*t, self.Draft)
            self.MassOb=Platform.add_mass_objects([self.MassOb, self.MassObMG])     
    
        def apply_special_floater(self, PlatformMass, DTList, DBList, HeightList, NSegmentList):
            #### Function that can generate any floater shape easily
            #### First get all the shapes 
            OutLengths=[]
            OutDiameters=[]
            SectionVol=[]
            for i, H in enumerate(HeightList):
                OutVol, OutOF = self.get_cone_properties(DTList[i], DBList[i], HeightList[i], NSegmentList[i])
                OutLengths.extend(OutOF[1])
                OutDiameters.extend(OutOF[0])
                SectionVol.extend(OutOF[2])
            
            #### Calc the draft
            CumVolInv=np.cumsum(np.flip(SectionVol))
            CumLenInv=np.cumsum(np.flip(OutLengths))
            
            self.DisplacedMassPlatform=PlatformMass+self.Mass*3
            self.Draft=np.interp(self.DisplacedMassPlatform/3/1.025, CumVolInv, CumLenInv)
            
            CumLen=np.cumsum(OutLengths)
            TotalLen=np.sum(OutLengths)
            
            #### split into wet and dry floater at that point (useful info to determine draft and later for MG calcs)
            wet=(TotalLen-CumLen)<self.Draft
            dry=(TotalLen-CumLen)>self.Draft
            OutDiameters=np.array(OutDiameters)
            OutLengths=np.array(OutLengths)
            Ones=np.ones(np.shape(OutDiameters))
            
            self.CylinderDiameter=np.concatenate([OutDiameters[dry],[OutDiameters[wet][0]],OutDiameters[wet]])
            self.CylinderLength=np.concatenate([OutLengths[dry],[TotalLen-CumLen[dry][-1]-self.Draft, self.Draft-(TotalLen-CumLen[wet][0])], OutLengths[wet][1:]])
            self.CdNormal=np.concatenate([[self.CdNormalAboveWL], Ones[dry]*self.CdNormalAboveWL,Ones[wet]*self.CdNormalUnderWL])                                     

            #### Drag area is determined using the one to last floater section
            self.AreaAxial=np.pi/4*(self.CylinderDiameter[-3])**2
            self.NumCylinders=len(self.CylinderDiameter)        
        
        def set_special_floater_inertia(self, Platform, Masses,Sections):
            self.MassOb=MassObject(0, [0,0,0])
            for iM, M in enumerate(Masses): 
                for S in Sections[iM]: 
                    L=np.sum(self.CylinderLength[S:])-self.CylinderLength[S]/2
                    TmpMassOb=MassObject(M/len(Sections[iM]), [0,0,L])
                    TmpMassOb.set_cylinder_inertia(self.CylinderDiameter[S], self.CylinderLength[S])
                    self.MassOb=Platform.add_mass_objects([self.MassOb, TmpMassOb])
                             
class DragPlate:
    def __init__(self, Diameter=2.25, PlateType=0, DragCoefficient=1.5):
        self.Type=PlateType
        self.Diameter=Diameter
        self.Cd=DragCoefficient
        self.Ca=0.64 ##Based on Ca of flat plate DNV c205
        
    def calc_details(self, Platform, Floater):
        self.Radius=self.Diameter/2
        self.Area=self.Radius**2*np.pi
        PlateReferenceVolume=4/3*(self.Diameter/2)**3*np.pi
        ### Check if platform and floater 
        # if not round(Floater.Draft*Floater.AreaAxial*1.025,3) == round(Platform.Mass/3,3):
        #     print('Floater draft and platform mass do not match')
        ### Plate type one is directyl at the end of the cylinder
        if self.Type==1:
            if self.Diameter-Floater.Diameter<0 :
                print('Unrealistic plate type 1, plate diameter smaller than floater diameter')
            CutOffHeight=np.sqrt(self.Radius**2-(Floater.Diameter/2)**2)
            VolumeCylinder=Floater.AreaAxial*CutOffHeight
            H=self.Radius-CutOffHeight
            VolumeCap=1/3*np.pi*H**2*(3*self.Radius-H)
            AddedVolume=PlateReferenceVolume-VolumeCylinder-VolumeCap
            self.HydrodynamicMass=AddedVolume*1.025
            self.Z=Floater.PosBottom[2]
            Floater.CaAxial=0
            Floater.CdAxial=0
        if self.Type==2:
            ## Plate type two is always applied at a depth such that added masses are independent 
            self.Z=Floater.PosBottom[2]+(self.Radius+Floater.Diameter/2)*-1
            self.HydrodynamicMass=PlateReferenceVolume*1.025
        ## Drag area for now is kept constant for both plates based on two consequtive plaves Cd DNV c205
        ## End of cylinder drag coefficient contribution is also added
        DragArea=self.Area-Floater.AreaAxial
        self.DragArea=DragArea+Floater.CdAxial/self.Cd*self.Area

class Coupling:
    
    def __init__(self):
        self.PosCenter=np.array([0, 21.25/2-2.312, 3.45]) ## Only used for Y and Z coordinates
        self.Gap= 3.715
        self.Height=3.4
        self.Width=2.09
        self.Stiffness={}
        self.Stiffness['In']=750
        self.Stiffness['Out']=740
        self.Stiffness['Mid']=100000
        self.PosGapProbe=self.PosCenter*[0,1,1]
        self.TypeOptions=['In', 'Out', 'Mid']
        self.Damping={}
        self.LinearSpring={}
        self.LinearDamping={}
        for Type in self.TypeOptions:
            self.Damping[Type]=0.1
            self.LinearSpring[Type]='Yes'
            self.LinearDamping[Type]='Yes'
        self.Loc=[-1,1]
        self.LoadCells=True
       
        # self.PosX_T1=-self.Gap
        self.PosX_T1=0
        self.PosX_T2=0
        self.dXBottomTop=0.213
        self.Alternating=False
        
    def calc_details_origional(self):
        self.Position={}
        self.Len={}
        self.Position['Out']= self.PosCenter+np.array([0,self.Width/2,0])
        self.Position['In']= self.PosCenter-np.array([0,self.Width/2,0])
        self.Position['Mid']= self.PosCenter-np.array([0,0, self.Height])
        # self.Position['Center']=self.PosCenter-[self.Gap, 0, 0] # This is the center position in the same coordinated as the other positions
        self.Position['Center']=self.PosCenter+[self.PosX_T1,0,0] # test
        self.Len['In']=np.sqrt(self.Gap**2+(self.Width/2)**2)
        self.Len['Out']=self.Len['In']
        self.Len['Mid']=np.sqrt(self.Gap**2+(self.Height)**2)   
        
    def calc_details(self):
        self.Position={}
        self.Len={}
        self.Position['Out']= self.PosCenter+[self.PosX_T2-self.dXBottomTop,self.Width/2,0]
        self.Position['In']= self.PosCenter+[self.PosX_T2-self.dXBottomTop,-self.Width/2,0]
        self.Position['Mid']= self.PosCenter+[self.PosX_T2,0, -self.Height]
        # self.Position['Center']=self.PosCenter-[self.Gap, 0, 0] # This is the center position in the same coordinated as the other positions
        self.Position['Center']=self.PosCenter+np.array([self.PosX_T1,0,0]) # test
        self.Len['In']=np.sqrt((self.Gap-self.dXBottomTop)**2+(self.Width/2)**2)
        self.Len['Out']=self.Len['In']
        self.Len['Mid']=np.sqrt(self.Gap**2+(self.Height)**2) 
        
    def calc_details_down_position(self,UnstretchedCylinderLength, CylinderStroke, ExtraDeckClearance=0):
        ## For this funtion coupling height and width should be correctly described
        ## Should be performed after calc details to keep the postions correct
        CylinderLengthIn=UnstretchedCylinderLength-CylinderStroke
        GapIn=(CylinderLengthIn**2-(self.Width/2)**2)**0.5
        self.GapMid=(UnstretchedCylinderLength**2-(self.Width/2)**2)**0.5
        self.Len['Mid']=((GapIn+self.dXBottomTop)**2+(self.Height-ExtraDeckClearance)**2)**0.5
        self.Len['In']=UnstretchedCylinderLength
        self.Len['Out']=self.Len['In']
        self.phi_opt()
        self.Gap=self.Len['Mid']*np.cos(self.Phi*np.pi/180)
        self.H2=self.Len['Mid']*np.sin(self.Phi*np.pi/180)
        self.Position['Center']=-np.array([0,0,(self.Height-self.H2)])+self.Position['Center']
        #### Line to place center location on the same x position of both platforms
        self.Position['Center'][0]=-self.Gap+self.PosX_T2
        
            
    def opt_fun(self,Phi):
        return (((self.Height-self.Len['Mid']*np.sin(Phi*np.pi/180))**2+(self.dXBottomTop-self.Len['Mid']*np.cos(Phi*np.pi/180))**2)**0.5-self.GapMid)**2
    
    def phi_opt(self):
        self.Phi=optimize.fmin(self.opt_fun,0,disp=False)[0]
        
    
    def calc_details_config10(self):
        self.LinearSpring['Mid']='No'
        self.LinearSpring['In']='Yes'
        self.LinearSpring['Out']='Yes'
        self.Stiffness['In']=100000
        self.Stiffness['Out']=100000
        self.Damping['In']=0
        self.Damping['Out']=0
        self.Position={}
        self.Len={}
        self.Position['Out']= self.PosCenter+[0,self.Width/2,-self.Height]
        self.Position['In']= self.PosCenter-[0, self.Width/2, self.Height]
        self.Position['Mid']= self.PosCenter-[0, 0, 0]
        # self.Position['Center']=self.PosCenter-[self.Gap, 0, 0] # This is the center position in the same coordinated as the other positions
        self.Position['Center']=self.PosCenter # test
        self.Len['In']=np.sqrt(self.Gap**2+(self.Width/2)**2+self.Height**2)
        self.Len['Out']=self.Len['In']
        self.Len['Mid']=self.Gap  
        self.TypeOptions=['In', 'Mid','Out']
        
    def calc_details_config10a(self):
        ## Only applicable after application of config 10 
        self.calc_details_config10()
        self.Position['In']=self.PosCenter-[0,0, self.Height]
        self.Len['In']=np.sqrt(self.Height**2+self.Gap**2)
        self.TypeOptions=['In', 'Mid']
            
    def get_array_excel(self, wb, sheet, loc):
        ws=wb[sheet]
        cells=ws[loc]
        Out=[]
        for i in cells[0]:
            Out.append(i.value)
        return np.array(Out)
    
    def set_non_linear_stiffness(self,wb, sheet):
        self.NLDisplacement=self.get_array_excel(wb, sheet, 'H50:AV50')
        self.NLForce=self.get_array_excel(wb, sheet, 'H62:AV62')
        self.LinearSpring['In']='No'
        self.LinearSpring['Out']='No'
        self.NLLengthZeroIntercept=np.interp(0, self.NLForce, self.NLDisplacement)
    
    def set_critical_damping(self, Platform, Floater, Ratio=1):
        ## Assumes In and Out stiffness are the same critical damping based on relative surge motions between platforms 
        k=self.Gap/self.Len['In']*self.Stiffness['In']*4
        m=(Platform.MassT1+Platform.MassT2)/2*(1+Floater.CaNormal)
        CriticalDampingTotal=2*np.sqrt(k*m)
        Damping=CriticalDampingTotal/4*self.Len['In']/self.Gap
        self.Damping['In']=Damping*Ratio
        self.Damping['Out']=Damping*Ratio
    
    def set_damping_power(self, Power):
        ## sets quadratic damping with the cylinder set damping at 1m/s
        self.NLDampingVelocity=np.linspace(0,3,70)
        self.NLDampingForce=self.Damping['In']*self.NLDampingVelocity**Power
        self.LinearDamping['In']='No'
        self.LinearDamping['Out']='No'
    
    def calc_names_and_locations(self, Layout, Platform, Compose):
        ## This function set all the names and positions for the coupling class, these details can then be extract at a later stage for composing the model and plotting and post processing
        Type=Layout.Matrix.flatten()
        self.Names=[]
        self.PosEndAs=[]
        self.PosEndBs=[]
        self.EndAConnections=[]
        self.EndBConnections=[]
        self.Types=[]
        self.Directions=[]
        # self.PosPlotGlobal=[]
        self.PlatformOrientations=[]
        self.PosProbeEndAs=[]
        self.PosProbeEndBs=[]
        self.Locs=[]            
        # self.PosPlotGlobalProbe=[]
        for i in range(len(Type)):
            if Type[i]==2:
                ind=np.array([True, True, True])
                # Determine if a copling is neccessary, first on posistion in grid, thereafter on adjacent platform
                if np.mod(i,Layout.Matrix.shape[1])==0:
                    ind[2]=False
                elif Type[i-1]==0:
                    ind[2]=False 
                if np.mod(i+1,Layout.Matrix.shape[1])==0:
                    ind[1]=False 
                elif Type[i+1]==0:
                    ind[1]=False
                if i>=Layout.Matrix.shape[1]*(Layout.Matrix.shape[0]-1):
                    ind[0]=False
                elif Type[i+Layout.Matrix.shape[1]]==0:
                    ind[0]=False 
                for j in range(3):
                    direc=[0,120,240]
                    if ind[j]:
                        for T in self.TypeOptions: ## Typically this will be in out mid 
                            for k in self.Loc: ## Loc determines on which side the coupling comes
                                ## Calculate all the coupling details for future reference
                                self.Names.append(T+ '_' + str(i)+'_'+str(direc[j])+'_'+str(k))
                                self.EndAConnections.append('Platform_'+str(i))
                                PosA=Compose.translate_coordinates(Platform,self.Position[T]*[1,k,1], direc[j])
                                PosProbeEndAs=Compose.translate_coordinates(Platform,self.PosGapProbe*[1,k,1], direc[j])
                                self.PosEndAs.append(PosA)
                                self.PosProbeEndAs.append(PosProbeEndAs)
                                if direc[j]==0:
                                     self.EndBConnections.append('Platform_'+str(i+Layout.Matrix.shape[1]))
                                if direc[j]==120:
                                     self.EndBConnections.append('Platform_'+str(i+1))
                                if direc[j]==240:
                                     self.EndBConnections.append('Platform_'+str(i-1))                                    
                                PosB=Compose.translate_coordinates(Platform, self.Position['Center']*[1,k*-1,1], direc[j]) 
                                PosProbeEndBs=Compose.translate_coordinates(Platform,self.PosGapProbe*[1,k*-1,1], direc[j])                                          
                                self.PosEndBs.append(PosB)
                                self.PosProbeEndBs.append(PosProbeEndBs)
                                self.Types.append(T)
                                self.Directions.append(direc[j])
                                self.PlatformOrientations.append((Type[i]-1)*180)
                                self.Locs.append(k)
        if Platform.Split:
            self.split_connections()
    
    def set_alternating_coupling(self, Compose, Platform): 
        for i,Name in enumerate(self.Names):
            if Name.find('-1')!=-1:
                ConnectionA0=self.EndAConnections[i]
                ConnectionB0=self.EndBConnections[i]
                self.EndAConnections[i]=ConnectionB0
                self.EndBConnections[i]=ConnectionA0
                self.PosEndAs[i]=Compose.translate_coordinates(Platform,self.Position[self.Types[i]], self.Directions[i])
                self.PosEndBs[i]=Compose.translate_coordinates(Platform,self.Position['Center']*[1,-1,1], self.Directions[i])
                # if Platform.Split: 
                    
        self.Alternating=True
        Platform.AlternatingCoupling=True
    
    def split_connections(self):
        for i in range(len(self.EndAConnections)):
            self.EndAConnections[i]=self.EndAConnections[i]+'_'+str(np.mod(self.Directions[i]+60-60*self.Locs[i],360))
            self.EndBConnections[i]=self.EndBConnections[i]+'_'+str(np.mod(self.Directions[i]+60+60*self.Locs[i],360))        
                                                        
 
    def scale_size(self, ScaleOrigin, ScaleNew):
        ## Be careful when using this scaling function in loops!!
        R=ScaleNew/ScaleOrigin
        self.PosCenter=self.PosCenter*R
        self.Gap=self.Gap*R
        self.Height=self.Height*R
        self.Width=self.Width*R
        for s in self.Stiffness.keys():
            self.Stiffness[s]=self.Stiffness[s]*R**2
            self.Damping[s]=self.Damping[s]**2.5
        self.PosGapProbe=self.PosGapProbe*R     

    def scale_stiffness(self, ScaleOrigional, ScaleNew):
        ## Be careful when using this scaling function in loops!!
        R=ScaleNew/ScaleOrigional
        self.NLForce=self.NLForce*R**2
        self.Stiffness['In']=self.Stiffness['In']*R**2
        self.Stiffness['Out']=self.Stiffness['Out']*R**2
        self.Stiffness['Mid']=self.Stiffness['Mid']*R**2
    
    def find_edged_platforms(self,matrix):
        NoneZeroRows = []
        NonZeroRowIndex = []
        NoneZeroPos = []
        PlatformNum = []
        ####Step1 find the none zero rows in matrix
        for i, row in enumerate(matrix):
            if any(row):
                NonZeroRowIndex.append(i)
                NoneZeroRows.append(row)
        ####Step2 find the positions of each none zero elemnt, positions are positions in the matrix
        for ir, nrow in enumerate(NoneZeroRows):
            NoneZeroPos.append([j + matrix.shape[1]*NonZeroRowIndex[ir] for j, val in enumerate(nrow) if val != 0])
        #### Step3 find the edge elements
        #### all the elements in the first and the last rows are accounted
        #### only the first two and the last two elements in the rest rows are accounted
        for ip, pfn in enumerate(NoneZeroPos):
            if ip == 0 or ip == (len(NoneZeroPos)-1):
                PlatformNum.append(pfn)
            else:
                PlatformNum.append([jj for iip, jj in enumerate(pfn) if iip == 0 or iip == 1 or iip == (len(pfn)-1) or iip == (len(pfn)-2)])
        return NoneZeroRows,NoneZeroPos,NonZeroRowIndex,PlatformNum
    
class Mooring: 
        def __init__(self, Platform, Coupling, Environment):
            self.Type='Tether'
            self.LengthChainBottom=20 #m
            self.LengthChainTop=5
            self.RadiusAnchor=175 #m
            self.LocationLineTypes=r"X:\01_Hydromechanics\Calculations\OrcaFlex\01_LineTypes\20230328_NylonUpdate.dat"
            self.M1LineTypes=['Chain32mm','Polyprop90mm','Chain32mm']
            # self.M2LineTypes=self.M1LineTypes[1:] # This for the Vlink mooring
            self.Position=Coupling.Position['Mid'].copy()#PosCenter.copy()*np.array([0,1,1])+[1.2,0,0]# This gives mooring connection position
            self.M1SegementLengths=[10, 70, 20] #This is only used for lines and not tethers
            self.VLinkLength=25 # Lenth for each part of the VLink 
            self.VLinkLineTypes=self.M1LineTypes[0]
            self.M2SegementLengths=[2]
            # self.M2SegementLengths=[2] #This is only used for lines and not tethers
            self.Depth=Environment.Depth
            self.ConnectedType=[1,2]
            self.ConnectedDirection=[0, 120, 240]
            self.MooringBuoy='No'
            self.CornerMooring=False 
            self.VLink=False
            self.Center=False
            self.AdjustBridalCornerAnchorPos = False
            self.RepositioningAnchorDegree = 5
            #### The properties below have been added to model the spring contents in Cork
            self.ContentsMethod='Free flooding'
            self.IncludeAxialContentsInertia='Yes'
        # Length ploy based on a straight line between the end of chain platform
            self.Angle=0 #### This sets up an outward (+) extra angle in the mooring lines
            self.FullStaticsConvergenceControlMethod='Line search'#'Mag. of std. error / change'
            self.StaticsStep2='None'
            self.StaticsSeabedFrictionPolicy='None'
            self.Representation='Finite element'  # 'Finite element' or 'Analytic catenary' method for mooring lines
        
        def calc_details(self, Floater, Layout):
            self.ConnectionAirgap=(-Floater.PosBottom[2]-Floater.Draft)+self.Position[2]
            self.LengthTotal=np.sqrt((self.ConnectionAirgap+self.Depth)**2+(self.RadiusAnchor)**2)
            self.LocLengths=[self.LengthChainTop, self.LengthTotal-self.LengthChainTop-self.LengthChainBottom, self.LengthChainBottom]
            self.VLinkMidLength=np.sqrt(np.sum(self.VLinkLength)**2-(Layout.EdgeLength/2-self.Position[1])**2)
            self.M2Lengths=[self.LengthTotal-self.LengthChainBottom-self.VLinkMidLength, self.LengthChainBottom]
            
        def calc_details_constant_synthline_length(self, Floater, Layout, SyntheticLineLength):
            #### This section calculate the mooring lengths, by keepin the first two section lengths (bottom chain and synthetic line lengths) constant 
            #### The top section length is adjusted for the correct waterdepth, draft etc. 
            self.ConnectionAirgap=(-Floater.PosBottom[2]-Floater.Draft)+self.Position[2]
            self.LengthTotal=np.sqrt((self.ConnectionAirgap+self.Depth)**2+(self.RadiusAnchor)**2)
            self.self.LocLengths=[self.LengthTotal-SyntheticLineLength-self.LengthChainBottom, SyntheticLineLength, self.LengthChainBottom]
            # print('Top chain length = '+str(round(self.M1Lengths[0],2))+'m')
            self.VLinkMidLength=np.sqrt(np.sum(self.VLinkLength)**2-(Layout.EdgeLength/2-self.Position[1])**2)
            self.M2Lengths=[self.LengthTotal-self.LengthChainBottom-self.VLinkMidLength, self.LengthChainBottom]
            self.SyntheticLineLength=SyntheticLineLength
            
        def set_top_chain_length(self, Floater, VLink, TotalLength='Automatic'):
            if TotalLength=='Automatic':
                #### This set the line length to the geometric length
                self.ConnectionAirgap=(-Floater.PosBottom[2]-Floater.Draft)+self.Position[2]
                self.LengthTotal=np.sqrt((self.ConnectionAirgap+self.Depth)**2+(self.RadiusAnchor)**2)
            else:
                self.LengthTotal=TotalLength
                # VLink = True
            if VLink:
                self.M2Lengths[0]=self.LengthTotal-np.sum(self.M2Lengths[1:])
                print('Top chain length = '+str(round(self.M2Lengths[0],2))+'m')
            else:
                self.M1Lengths[0]=self.LengthTotal-np.sum(self.M1Lengths[1:])
                print('Top chain length = '+str(round(self.M1Lengths[0],2))+'m')
        
        # def calc_vlink_mooring_details(self,Floater,VLinkLength):
        #     self.VLinkLength = VLinkLength
        #     self.ConnectionAirgap=(-Floater.PosBottom[2]-Floater.Draft)+self.Position[2]
        #     self.MooringAngle = np.arctan((self.ConnectionAirgap + self.Depth) / self.RadiusAnchor) *180/np.pi ### in degree
        #     self.VLinkMidLength = (np.sum(self.VLinkLength)**2 - self.Position[1]**2)**0.5
        #     self.VLinkMidHoriLength = np.sum(self.VLinkLength) * np.cos(np.radians(self.MooringAngle))
        #     self.VLinkBuoyPositon = np.array([self.Position[0]-self.VLinkMidHoriLength,0,self.Position[2] - self.VLinkMidLength*np.sin(np.radians(self.MooringAngle))])
        
        def calc_vlink_mooring_details(self,Floater,VLinkMidLength):
            self.VLinkMidLength = VLinkMidLength
            self.VLinkLength = (self.VLinkMidLength**2+self.Position[1]**2)**0.5
            self.ConnectionAirgap=(-Floater.PosBottom[2]-Floater.Draft)+self.Position[2]
            self.MooringAngle = np.arctan((self.ConnectionAirgap + self.Depth) / self.RadiusAnchor) *180/np.pi ### in degree
            self.VLinkMidHoriLength = self.VLinkMidLength * np.cos(np.radians(self.MooringAngle))
            self.VLinkBuoyPositon = np.array([self.Position[0]-self.VLinkMidHoriLength,0,self.Position[2] - self.VLinkMidLength*np.sin(np.radians(self.MooringAngle))])
            self.VlinkMidAnchorPos = np.array([self.Position[0]-self.RadiusAnchor,0,-self.Depth])
            self.VlinkCornerAnchorPos1 = np.array([self.Position[0]-self.RadiusAnchor*(1-np.deg2rad(self.RepositioningAnchorDegree)*np.cos(np.deg2rad((180-self.RepositioningAnchorDegree)/2))),
                                                self.RadiusAnchor*np.deg2rad(self.RepositioningAnchorDegree)*np.sin(np.deg2rad((180-self.RepositioningAnchorDegree)/2)),
                                                -self.Depth])
            self.VlinkCornerAnchorPos2 = np.array([self.Position[0]-self.RadiusAnchor*(1-np.deg2rad(self.RepositioningAnchorDegree)*np.cos(np.deg2rad((180-self.RepositioningAnchorDegree)/2))),
                                                -self.RadiusAnchor*np.deg2rad(self.RepositioningAnchorDegree)*np.sin(np.deg2rad((180-self.RepositioningAnchorDegree)/2)),
                                                -self.Depth])
        def calc_connected_platforms(self, Layout, Directions=[0,120,240]):
            Type=Layout.Matrix.flatten()
            self.ConnectedPlatformsTotal=[]
            self.OrientationLocalTotal=[]
            self.ConnectedPlatformType=[]
            self.LocTotal=[]
            ## Note there will be problems with Loc if a platforms is moored on two side, this is however very unlikely to occur
            for i, t in enumerate(Type):
                ind = [False, False, False] 
                ## top and bottom boundary
                if t==1 and i< Layout.Matrix.shape[1]: ### First row
                    ind[0]= True
                    Loc=1
                if t==1 and i> Layout.Matrix.shape[1]:
                    if Type[i-Layout.Matrix.shape[1]]==0:
                        ind[0]=True
                        Loc=1
                if t==2 and i>Layout.Matrix.shape[1]*(Layout.Matrix.shape[0]-1):
                    ind[0]= True
                    Loc=-1
                if t==2 and i< Layout.Matrix.shape[1]*(Layout.Matrix.shape[0]-1):
                    if Type[i+Layout.Matrix.shape[1]]==0:
                        ind[0]=True
                        Loc=-1
                ## side boundaries on the right
                if t==1 and np.mod(i,Layout.Matrix.shape[1])==0:
                    ind[1]=True
                    Loc=-1
                if t==1 and np.mod(i,Layout.Matrix.shape[1])!=0:
                    if Type[i-1]==0:
                        ind[1]=True
                        Loc=-1
                if t==2 and np.mod(i,Layout.Matrix.shape[1])==0:
                    ind[2]=True
                    Loc=-1
                if t==2 and np.mod(i,Layout.Matrix.shape[1])!=0:
                    if Type[i-1]==0:
                        ind[2]=True
                        Loc=-1
                ## side boundaries on the left 
                if t==1 and np.mod(i+1,Layout.Matrix.shape[1])==0:
                    ind[2]=True
                    Loc=1
                if t==1 and np.mod(i+1,Layout.Matrix.shape[1])!=0:
                    if Type[i+1]==0:
                        ind[2]=True
                        Loc=1
                if t==2 and np.mod(i+1,Layout.Matrix.shape[1])==0:
                    ind[1]=True
                    Loc=1
                if t==2 and np.mod(i+1,Layout.Matrix.shape[1])!=0:
                    if Type[i+1]==0:
                        ind[1]=True
                        Loc=1
                for j, direc in enumerate(Directions):
                    if ind[j]:
                        self.ConnectedPlatformsTotal.append(i)
                        self.OrientationLocalTotal.append(direc)
                        self.ConnectedPlatformType.append(t)    
                        self.LocTotal.append(Loc)
                        
class MooringBuoy:
    def __init__(self, Mooring):
        self.NetBuoyancy=100/9.81 ## Ton
        # self.AspectRatio=1  ## Height/Diamter
        self.MassRatio=0.2 ## Mass wrt NetBuoyancy
        self.CaSurge=0.6
        self.CdHeave=0.6
        # self.MooringRadius=20 ## radius from platform #### This is used for the old mooring buoy function as distance the buoy has from the platform
        self.BuoyLocation=15+33.15 #### Distance of mooring buoy location from side A or side B
        self.ReferenceEnd='End B'#### Side B is the anchor point
        Mooring.MooringBuoy='Yes'
        # self.calc_details()
        self.Offset=0 ## This is the offet on the mooring line
        self.MG='NMG'
    
    def calc_details_old(self):
        #### Function belonging to old mooirng buoy, not 
        self.Volume=self.NetBuoyancy*(1+self.MassRatio) /1.025  
        self.Mass=self.NetBuoyancy*self.MassRatio
        self.Radius=(self.Volume/(2*self.AspectRatio*np.pi))**(1/3)
        self.Height=self.AspectRatio*2*self.Radius
        self.DragAreaHeave=self.Radius**2*np.pi
        self.DragAreaSurge=self.Radius*2*self.Height/2 #Assume buoy is half submerged 
        self.CaHeave=4/3*np.pi*self.Radius**3*0.64/2*1.025 / (self.Volume/2*1.025) # Based on half flat plate and buoy half submerged
        
    def calc_details_sphere_buoyant(self):
        #### Calcs sphere details based on the net buoyancy and the mass ratio
        self.Volume=self.NetBuoyancy*(1+self.MassRatio) /1.025  
        self.Mass=self.NetBuoyancy*self.MassRatio
        self.Radius=(self.Volume*3/(4*np.pi))**(1/3)
        self.Height=self.Radius*2
        self.DragAreaHeave=self.Radius**2*np.pi
        self.DragAreaSurge=self.Radius**2*np.pi #Assume buoy is sumbmerged
        self.CaSurge=0.5
        self.CaHeave=0.5
        self.CdSurge=0.5
        self.CdHeave=0.5
        
    def calc_details_cylinder_buoyant(self):
        #### Calcs cylinder details based on the net buoyancy and the mass ratio
        self.Volume = self.NetBuoyancy*(1+self.MassRatio) /1.025  
        self.Mass = self.NetBuoyancy*self.MassRatio
        self.Height = self.Height
        # self.Radius = self.Radius
        self.Radius = (self.Volume/(self.Height*np.pi))**0.5
        self.DragAreaHeave = self.Radius**2*np.pi
        self.DragAreaSurge = self.Radius*2*self.Height
        self.CaSurge=0.5
        self.CaHeave=0.5
        self.CdSurge=0.5
        self.CdHeave=0.5        
        
    def calc_details_sphere_specificgravity(self):
        #### Calcs sphere details based on the net buoyancy and the mass ratio
        self.Volume=self.Mass/self.SpecificGravity/1.025
        # self.Mass=self.NetBuoyancy*self.MassRatio
        self.Radius=(self.Volume*3/(4*np.pi))**(1/3)
        self.Height=self.Radius*2
        self.DragAreaHeave=self.Radius**2*np.pi
        self.DragAreaSurge=self.Radius**2*np.pi #Assume buoy is sumbmerged
        self.CaSurge=0.5
        self.CaHeave=0.5
        self.CdSurge=0.5
        self.CdHeave=0.5
    
    def add_marine_growth(self, Thickness, Density):
        VolumeMG=4/3*np.pi*((self.Radius+Thickness)**3-self.Radius**3)
        self.Mass=self.Mass+VolumeMG*Density
        self.Radius=self.Radius+Thickness
        self.Volume=self.Volume+VolumeMG
        self.Height=self.Radius*2
        self.DragAreaHeave=self.Radius**2*np.pi
        self.DragAreaSurge=self.Radius**2*np.pi #Assume buoy is sumbmerged
        self.MG='MG'
        
    def add_marine_growth_cylindershape(self, Thickness, Density):
        VolumeMG = np.pi*((self.Radius+Thickness)**2-self.Radius**2)*self.Height ### 4/3*np.pi*((self.Radius+Thickness)**3-self.Radius**3)
        self.Mass=self.Mass+VolumeMG*Density
        self.Radius=self.Radius+Thickness
        self.Volume=self.Volume+VolumeMG
        self.Height=self.Height###self.Radius*2
        self.DragAreaHeave=self.Radius**2*np.pi
        self.DragAreaSurge=self.Radius*2*self.Height ###self.Radius**2*np.pi #Assume buoy is sumbmerged
        self.MG='MG'                       
        
    def calc_line_lengths(self, Mooring, Floater, Layout):
        ## Calc Vlink length, M1a length
        Mooring.VLinkLength=(self.MooringRadius**2+(Layout.EdgeLength/2-Mooring.Position[1])**2+Mooring.ConnectionAirgap**2)**0.5
        Mooring.M1aLengths=[(self.MooringRadius**2+Mooring.ConnectionAirgap**2)**0.5]
        Mooring.calc_details(Floater,Layout)
        Mooring.M1aLineTypes=['Polyprop90mm']
        Mooring.M1aSegementLengths=[30]
    
    def set_name(self):
        self.Name='MooringBuoy_'+str(round(self.NetBuoyancy,2))+'t_'+ self.MG
        
class Layout:
    ## Input
        def __init__(self):
            # self.Matrix=np.array([[2,1,2,1,2,1,2],[1,2,1,2,1,2,1],[0,1,2,1,2,1,0],[0,0,1,2,1,0,0]]) #5 sided triagle set-up 
            self.Matrix=np.array([[2,1,2,1,2],[1,2,1,2,1],[0,1,2,1,0]]) #4 sided triangle set-up
            self.MatrixDefinition='Rows and collums of the layout, 0 indicates no platform, 1 a platform with a center connection, 2 a platform with 3 connections'
            self.Flipped=False
            
        def calc_matrix_hexagonal(self, n):
        ##  n is the edgelength of the hexagon
            Mat=np.matrix([[2,1,2],[1,2,1]])
            AppMat=np.matrix([[1,2,1,2],[2,1,2,1]])
            
            if n > 1: 
                for r in range(int(n)-1):
                    Mat=np.append(Mat,AppMat,1)
                Mat=np.matlib.repmat(Mat,n,1)
                if (n%2)==0:
                    Mat=np.flip(Mat)
                for r in np.arange(n-1):
                        Mat[r,:(n-r-1)]=0            
                        Mat[r,-1*(n-r-1):]=0
                        Mat[-1*(r+1),:(n-r-1)]=0
                        Mat[-1*(r+1),-1*(n-r-1):]=0
            self.Matrix=np.array(Mat)
            
        def calc_matrix_snake(self, Width, Length):
            #Width is in terms of rows of 2
            #Length is every extra length more of the base length
            Mat=np.matrix([[1,2,1],[2,1,2]])
            AppMat=np.matrix([[2,1],[1,2]])
            
            for w in range(Width):
                Mat=np.matlib.repmat(Mat, w+1,1)
                AppMat=np.matlib.repmat(AppMat, w+1,1)
            for L in range(Length-1):
                Mat=np.append(Mat, AppMat,1)
            for w in np.arange(Width-1):
                Mat[w,:(Width-w-1)]=0
                Mat[w,-1*(Width-w-1):]=0
                Mat[-1*(w+1),:(Width-w-1)]=0
                Mat[-1*(w+1),-1*(Width-w-1):]=0
            self.Matrix=np.array(Mat)
       
        def calc_triangle(self, Edges, CutOffEdges):
            ## Create a triangle and with size edges
            ## Cuts of a traigle with size CutoffEdges
            
            ## Make the basic Matrix
            Mat=np.matrix([[1]])
            AppMat=np.matrix([[2,1]])
            
            for w in range(Edges-1):
                Mat=np.append(Mat, AppMat,1)
            Mat2=Mat
            for w in range(Edges-1):
                if np.mod(w,2)==0:
                    AppMat1=np.mod(Mat,2)+1    
                    Mat2=np.append(Mat2, AppMat1,0)
                else:
                    Mat2=np.append(Mat2, Mat,0)
            ## Add zeros to make the triangle 
            for r in np.arange(Edges-1):
                Mat2[Edges-r-1,:(Edges-r-1)]=0            
                Mat2[Edges-r-1,-1*(Edges-r-1):]=0
                
            ## Cut off triangle
            for r in np.arange(CutOffEdges):
                ## Center
                Mat2[Edges-r-1,:]=0
                ##Bottom Left
                Mat2[r,:(CutOffEdges-1)*2+1-r]=0    
                ##Bottom right
                Mat2[r,-1*((CutOffEdges-1)*2-r+1):]=0
                # Mat[-1*(r+1),-1*(n-r-1):]=0
            self.Matrix=np.array(Mat2)       
            
        def calc_matrix_christmastree(self, Order=3):
            MatOut=np.zeros([3*Order, 3+(Order-1)*4])
            range1=np.flip(np.arange(Order))
            for i in range1:
                self.calc_triangle(3+i,0)
                Mat=np.array(self.Matrix)
                print(Mat)
                MatOut[Order*3-(i+1)*3:Order*3-(i+1)*3+Mat.shape[0],(Order-i):(Order-i)+Mat.shape[1]]=Mat
            self.Matrix=MatOut
        
        def calc_slender_shape(self,m,n,Orient1=True):
            #### m is number of triangles in the long side
            #### n is rows of short side
            #### Orient1 is the rows3 model with the shorter side of the two long sides facing the main coming waves
            ## Make the basic Matrix
            if Orient1:
                Mat=np.matrix([[2]])
                AppMat=np.matrix([[1,2]])
            else:
                Mat=np.matrix([[1]])
                AppMat=np.matrix([[2,1]])
              
            for w in range(m):
                Mat=np.append(Mat, AppMat,1)
            Mat2=Mat
            for w in range(n):
                if np.mod(w,2)==0:
                    AppMat1=np.mod(Mat,2)+1    
                    Mat2=np.append(Mat2, AppMat1,0)
                else:
                    Mat2=np.append(Mat2, Mat,0)
            ## Add zeros to make the shape
            for r in np.arange(Mat2.shape[0]):
                # print(r)
                if n == 2:
                    if r > n-1:
                        Mat2[r:,:] = 0
                    else:
                        Mat2[r,:m//2] = 0
                        Mat2[r,m//2+m:] = 0
                if n == 3:
                    if Orient1:
                        if r<= 1:
                            Mat2[r,:m//2] = 0
                            Mat2[r,m//2+m:] = 0
                        elif r > n-1:
                            Mat2[r:,:] = 0
                        else:
                            Mat2[r,:m//2+1] = 0
                            Mat2[r,m//2-1+m:] = 0
                    else:
                        if r == 0:
                            Mat2[r,:m//2+1] = 0
                            Mat2[r,m//2-1+m:] = 0
                        elif r > n-1:
                            Mat2[r:,:] = 0
                        else:
                            Mat2[r,:m//2] = 0
                            Mat2[r,m//2+m:] = 0
            self.Matrix=np.array(Mat2)
        
        def Noble_model_shape(self,m,n):
            #### m is number of triangles in the long side
            #### n is rows of short side
            ## Make the basic Matrix
            Mat=np.matrix([[2]])
            AppMat=np.matrix([[1,2]])
              
            for w in range(m):
                Mat=np.append(Mat, AppMat,1)
            Mat2=Mat
            for w in range(n):
                if np.mod(w,2)==0:
                    AppMat1=np.mod(Mat,2)+1    
                    Mat2=np.append(Mat2, AppMat1,0)
                else:
                    Mat2=np.append(Mat2, Mat,0)
            ## Add zeros to make the shape
            for r in np.arange(Mat2.shape[0]):
                if r == 0:
                    num = 1
                    Mat2[r,:m//2+num] = 0
                    Mat2[r,m//2-num+m:] = 0
                elif r > n-1:
                    Mat2[r:,:] = 0
                else:
                    Mat2[r,:m//2+num-1] = 0
                    Mat2[r,m//2-num+1+m:] = 0
            self.Matrix=np.array(Mat2)
        
        def Teel_model_shape(self):
            Mat2=np.matrix([[2]])
            self.Matrix=np.array(Mat2)
            
            
        def flip_matrix(self):
            ## This can be useful for triangular structures
            Mat=np.flipud(self.Matrix)
            Mat1=Mat.copy()
            Mat1[Mat==1]=2
            Mat1[Mat==2]=1
            self.Matrix=Mat1
            self.Flipped=True
    
        def calc_details(self, Platform, Coupling):
            ## Calc virtual lengths    
            self.Gap= Coupling.Gap-Coupling.Position['Mid'][0]-Coupling.Position['Center'][0]
            self.EdgeLength= Platform.EdgeLength + (self.Gap/2)*np.sqrt(3)*2 
            self.Length= self.EdgeLength*np.sqrt(3)/2
        
            ## Calc Y locations platform origins
            PlatformsY= np.arange(1,self.Matrix.shape[1]+1)*self.EdgeLength/2
            PlatformsY= PlatformsY-((0.5+self.Matrix.shape[1]*0.5)*self.EdgeLength)/2 #Cetralise platforms
            self.PlatformsY= np.matlib.repmat(PlatformsY,self.Matrix.shape[0], 1)# Repeat array for amount of rows
            self.PlatformsY=self.PlatformsY.flatten()
            ## 
            PlatformsX= np.arange(self.Matrix.shape[0])-1 + self.Matrix.transpose() #Create node matrix 
            self.PlatformsX= PlatformsX.transpose()*self.Length 
            self.PlatformsX[self.Matrix==1]=self.PlatformsX[self.Matrix==1]+self.Gap/2
            self.PlatformsX[self.Matrix==2]=self.PlatformsX[self.Matrix==2]-self.Gap/2
            self.PlatformsX=self.PlatformsX.flatten()
            
            ## Calc orientations
            self.Orientation=np.zeros(self.Matrix.shape)
            self.Orientation[self.Matrix==1]=0
            self.Orientation[self.Matrix==2]=180            
            self.Orientation=self.Orientation.flatten()
            self.Types=self.Matrix.flatten()
            
        def NumberPlatform(self):
            # check total number of platforms
            return np.sum(np.where(self.Matrix == 2, 1, self.Matrix))
class Wing: 
            
    def __init__(self, Platform):
        ### 83 percent of the plaform area is used as wing 
        self.Area=Platform.Length*Platform.EdgeLength/2*0.83 
        self.Chord=np.sqrt(self.Area)
        self.Span=self.Chord
        self.Azimuth=-90  ## positioning parameters
        self.Declination=90 ## positioning parameters
        self.Gamma=90 ## positioning parameters
        # LocWingCoefficients=r"C:\Users\Sander\PW2\Engineering - ENG\Hydromechanics\Calculations\Excel\20220204 WindLoadingNotes.xlsx"
        # wbWingCoefficients=load_workbook(LocWingCoefficients, data_only=True)
        # wsWingCoefficients=wbWingCoefficients['OrcaFlexInput']
        # Coefficients=wsWingCoefficients['A11':'P43'] 
        # Out2=[]
        # for i in range(len(Coefficients)):
        #     Out=[]
        #     for j in Coefficients[i]: 
        #         try:
        #             Out.append(float(j.value))
        #         except:
        #             Out.append(j.value)
        #     Out2.append(Out)
        # self.Coefficients=np.array(Out2)
        # self.interp_cm()
        #### This is prepared such coefficients can be placed row by row
        #### For inital versio non R1 coefficients will be used, R2 coefficients are less useful    
        self.WingTypes=['R1 Out', 'R1 In', 'R2 Out', 'R2 In']
        self.Cl={}
        self.Cd={}
        self.Cm={}
        self.Angles=self.Coefficients[:,0]
        for i, Type in enumerate(self.WingTypes):
            self.Cl[Type]=self.Coefficients[:,3*i+4]
            self.Cd[Type]=self.Coefficients[:,3*i+5]
            self.Cm[Type]=self.Coefficients[:,3*i+6]
        self.ConnectionPosition=np.array(Platform.CoG)*[1,1,0]+[0,0,1.7]
        self.Colour={}
        self.Colour['R1 Out']=10485760
        self.Colour['R1 In']= 16711680
        self.Colour['R2 Out']=12632256
        self.Colour['R2 In']= 12632256
        self.RowReductionFactors=[1,1]
        
    def interp_cm(self):
        ## funcation to linearly interp the missing Cm values
        for i in [6,9,12,15]:
            Cm=self.Coefficients[:,i]   
            AnglesCm=np.array(self.Coefficients[Cm!=None,0],dtype='float64')
            Cm=Cm[Cm!=None]
            Cm=np.array(Cm,dtype='float64')
            AnglesTotal=np.array(self.Coefficients[:,0],dtype='float64')
            CmTotal=np.interp(AnglesTotal,AnglesCm, Cm)
            self.Coefficients[:,i]=CmTotal
            
class ReferenceConstraint:
    def __init__(self, ConnectionPosition, GlobalOrientation):
        self.Position=ConnectionPosition ### Location where the contraint is connected to the platform 
        self.Orientation=GlobalOrientation

class Environment: 
    
    def __init__(self, Hs=2, Depth=20, WaveHeading=0, CurrentSpeed=0):
        self.Depth=Depth
        self.WaveHeading=WaveHeading
        self.CurrentSpeed=CurrentSpeed
        self.Hs=Hs
        self.Hmax=4
        self.DurationStatistics=3*3600# Total duration time for sea-state
        self.WindSpeed=0
        self.WindDirection=0
        self.SpreadingCoefficient=None
        self.NumberOfWaveDirections=7
        self.NumberOfWaveComponentPerDirection=75
        self.WaveStretching='Vertical stretching'
        self.CurrentDirection=self.WaveHeading
        self.HsOrHmax='Hs'
        self.OrcaWavePeriods=np.concatenate([np.arange(2.5,7,0.25), np.arange(7,15,0.5)])
        self.CurrentPowerLawExponent=7
  
    def set_gamma(self): # Set gamma based on Hs and Tp
        if self.Tp/np.sqrt(self.Hs) <= 3.6:
                self.Gamma=5
        elif self.Tp/np.sqrt(self.Hs) < 5:
                self.Gamma=np.exp(5.75-1.15*self.Tp/np.sqrt(self.Hs))
        else:
                self.Gamma=1
                
    def name(self):
        if self.HsOrHmax=='Hs':
            self.Name= ("_Wdp%.1f" % (self.Depth) +
                    "_Hs%.2f" % (self.Hs) +
                    "_Tp%.2f" % (self.Tp) + 
                    "_Wdr%.0f" % (self.WaveHeading) +
                    "_Cv%.2f" % (self.CurrentSpeed) + 
                    "_Cdr%.0f" % (self.CurrentDirection)+
                    "_Wnv%.2f" % (self.WindSpeed) + 
                    "_Wndr%.0f" % (self.WindDirection))
        elif self.HsOrHmax=='Hmax':
            self.Name= ("_Wdp%.1f" % (self.Depth) +
                    "_Hmx%.2f" % (self.Hmax) +
                    "_T%.2f" % (self.Tass) + 
                    "_Wdr%.0f" % (self.WaveHeading) +
                    "_Cv%.2f" % (self.CurrentSpeed) + 
                    "_Cdr%.0f" % (self.CurrentDirection)+
                    "_Wnv%.2f" % (self.WindSpeed) + 
                    "_Wndr%.0f" % (self.WindDirection))
        
    def wave_number_fenton(self):
        omega=2*np.pi/self.Tass
        alpha=omega**2*self.Depth/9.81
        beta=alpha*(np.tanh(alpha))**-0.5
        kd_fenton=(alpha+beta**2*np.cosh(beta)**-2)/(np.tanh(beta)+beta*(np.cosh(beta)**-2))       
        self.KFenton=kd_fenton/self.Depth
    
    # def opt_fun(self,k):
    #     omega=2*np.pi/self.Tp
    #     return (9.81*k*np.tanh(k*self.Depth)-omega**2)**2
    
    # def wave_number_opt(self):
    #     self.k=optimize.fmin(self.opt_fun,0,disp=False)
    
    def calc_spectral_details_jonswap(self):
        ## for this Tp and Hs should always be defined
        self.set_gamma()
        self.Tz=(0.6673 + 0.05037*self.Gamma- 0.006230*self.Gamma**2+0.0003341*self.Gamma**3)*self.Tp #Relation TpTz JONSWAP
        self.T1=(0.7303 + 0.04946*self.Gamma - 0.006556*self.Gamma**2+ 0.0003610*self.Gamma**3)*self.Tp
        
    def calc_Hmax_rayleigh(self, k=1):
        self.calc_spectral_details_jonswap()
        self.Hmax=k*self.Hs*np.sqrt(np.log(self.DurationStatistics/self.Tz)/2) 
        
    def calc_details_Hmax_Hb(self):
        self.wave_number_fenton()
        self.lambda_Tass=2*np.pi/(self.KFenton)
        self.Hb=0.88/self.KFenton*np.tanh(0.89*self.KFenton*self.Depth) ## Miche breaking limit
        # self.Hb=self.lambda_Tass*0.142*np.tanh(2*np.pi*self.Depth/self.lambda_Tass) ## Alt DNVGL breaking limit
        if self.Hmax>self.Hb:
            print('Hmax is above breaking limit')
        self.UrHb=self.Hb*self.lambda_Tass**2/self.Depth**3
        self.UrHmax=self.Hmax*self.lambda_Tass**2/self.Depth**3
        
    def calc_details_steep_seastate(self): # Calculated detailed spectral parameters
    #### Calculates details for the maximum wave steepness
        self.Sp=1/15 #Start value
        Tp0=np.sqrt(self.Hs*2*np.pi/(self.Sp*9.81))
        self.Tp=Tp0
        TpError=1
        while TpError>0.02:
            if self.Tp > 15:
                self.Sp=1/25
            elif self.Tp>8:
                self.Sp= np.interp(self.Tp, [8,15], [1/15,1/25])
            self.Tp=np.sqrt(self.Hs*2*np.pi/(self.Sp*9.81))
            TpError=self.Tp-Tp0
            Tp0=self.Tp.copy()
            # print(TpError)
        self.calc_spectral_details_jonswap()
        self.ShallowWaterHmax=0.78*self.Depth
        
    def calc_Tass(self, H=False): 
        ## Calculates that associated equal propability Tass based on DNVGL C.3.3 design wave method 
        ## Calculated value based on Hmax if not otherwise given
        self.calc_spectral_details_jonswap()
        v=(self.T1/self.Tz-1)**0.5
        C1=1+v**2/(1+v**2)**(3/2)
        C2=0.5*v/(1+v**2)
        self.Tass=self.T1*C1
        D=1
        if H:
            if H > self.Hmax:
                print('H > Hmax calc not possible')
            else: 
                ah=(0.4013+0.0131*self.Gamma+0.0012*self.Gamma**2+ 0.00005*self.Gamma**3)
                D=C2*self.Hs/H*(2*(self.Hmax**2-H**2/(ah*self.Hs)**2)-2*np.log(self.Hmax/H))**2
        self.TassMin=self.Tass-D
        self.TassMax=self.Tass+D
        
        
    def calc_TpHs_from_TassHmax(self, k=1, TpTass=True):
        ## This functions estimates the Tp from the Tass and Hmaz that corresponds to the design wave
        ## k could be set to 0.9 for extreme events acording to some sources, 1 is more common to use
        self.HsOrHmax='Hs'
        GamErr=1
        while GamErr>.01:
            if TpTass:
                #### This is easy to explain more common option
                self.Tp=self.Tass/0.9
            else:
                #### This is the DNVGL methodology reverse engineerd
                v=(((0.7303 + 0.04946*self.Gamma - 0.006556*self.Gamma**2+ 0.0003610*self.Gamma**3)/(0.6673 + 0.05037*self.Gamma- 0.006230*self.Gamma**2+0.0003341*self.Gamma**3))-1)**0.5
                C1=1+v**2/(1+v**2)**(3/2)
                T1=self.Tass/C1
                self.Tp=T1/(0.7303 + 0.04946*self.Gamma - 0.006556*self.Gamma**2+ 0.0003610*self.Gamma**3)
            self.Tz=(0.6673 + 0.05037*self.Gamma- 0.006230*self.Gamma**2+0.0003341*self.Gamma**3)*self.Tp
            self.Hs=self.Hmax/(k*0.5*np.log(self.DurationStatistics/self.Tz))**0.5
            Gam=self.Gamma
            self.set_gamma()
            GamErr=abs(Gam-self.Gamma)
            # print('1')
    
    
    # def calc_spectrum_JONSWAP(self):
    #     ## This was created to calculate wave spreading and the JONSWAP spectrum 
    #     ## At later stage noticed this functionality in OrcaFlex and thus this was not needed anymore
    #     ## Kept here for potential use in the future for analytical methods
    #     self.Omegas=np.arange(0.1, 3, 0.05)
    #     self.set_gamma()
    #     self.OmegaPeak=2*np.pi/self.Tp
    #     Sigmas=0.07*(self.Omegas<=self.OmegaPeak) + 0.09*(self.Omegas>self.OmegaPeak);
    #     A = np.exp(-((self.Omegas/self.OmegaPeak-1)/(Sigmas*np.sqrt(2)))**2)
    #     self.SpecDensity = ((1-0.287*np.log(self.Gamma)) * 5/16 * self.Hs**2 * self.OmegaPeak**4 * self.Omegas**(-5) * np.exp(-5/4 * (self.Omegas /self.OmegaPeak)**(-4))* self.Gamma**A)
    #     self.SpreadingDirections=np.mod(np.arange(self.WaveHeading-90, self.WaveHeading+91, 15),360)
    #     n=self.SpreadingCoefficient
    #     SpreadingDirRad=self.SpreadingDirections*np.pi/180
    #     DirMain=self.WaveHeading*np.pi/180
    #     Ds=[]
    #     for i, d in enumerate(SpreadingDirRad):
    #         Ds.append(math.gamma(1+n/2)/(np.sqrt(np.pi)*math.gamma(1/2+n/2))*np.cos(d-DirMain)**n)
    #     self.Ds=1/np.sum(Ds)*np.array(Ds)
      
    def select_wave_type(self):
        if self.UrHmax>50:
            WType="Cnoidal"
        elif self.UrHmax<15:
            WType="Stokes' 5th"
        else:
            WType='Dean stream'
        self.WaveType=WType
        
class CheckSummary:

    warnings.filterwarnings("ignore", message="overflow encountered in multiply")        
    Report_Count = 0
    
    def __init__(self, filename, LocSims, model=False, Chk_Summary=True, model_save=True):
        self.LocSims  = LocSims
        self.FileName = filename
        if not model:
            self.CheckModel = ofx.Model(os.path.join(self.LocSims, self.FileName))
        else:
            self.CheckModel = model
        #
        self.Objects = np.array(self.CheckModel.objects)
        #
        self.findPlatformNum()
        self.minT1   = str(min(self.T1PlatformNum))
        self.minT2   = str(min(self.T2PlatformNum))
        self.minMoor = str(min(self.MoorPlatformNum))
        self.minPlt  = str(min(self.PlatformsNum))
        #
        if not str('G_Platform_' + self.minT1) in self.GroupNames:
            self.CreateGroups(self.PlatformsNum) # and save model
        if model_save:
            self.Anchor_Allowables(allowableGX = 1000, allowableGZ = 150)
            self.ConstraintHidden()
            self.CheckModel.SaveData(os.path.join(self.LocSims, self.FileName))
            print('Saved: ', self.FileName)
            self.CollapsedGroups(self.PlatformsNum) 
        #
        if Chk_Summary:            
            CheckSummary.Report_Count +=1
            print('CheckSummary.Report_Count = ',CheckSummary.Report_Count)
            self.SheetName = self.FileName.split("_")[3]
            self.SheetName = str(CheckSummary.Report_Count)+"-"+self.SheetName.split("-")[-1] # to reduce length
            self.ExcelFile = os.path.join(self.LocSims,'01_ModelCheckSummmary.xlsx')
            if os.path.isfile(self.ExcelFile) and CheckSummary.Report_Count==1:
                os.remove(self.ExcelFile)

            self.FloaterObj1     = 'Floater'   + self.minT2 + '_0'
            self.FloaterObj2     = 'Platform_' + self.minT2 + '_0'

            self.WaterDepth = self.CheckModel.environment.WaterDepth
            self.FloaterTopHeight = np.max(self.CheckModel[self.FloaterObj2].VertexZ[:])
            self.FloaterBottomHeight = self.CheckModel[self.FloaterObj1].CumulativeCylinderLength[-1]-self.FloaterTopHeight
    
            self.FloaterCLength = self.CheckModel[self.FloaterObj1].CumulativeCylinderLength[:]
            for i, CLen in enumerate(self.FloaterCLength):
                if CLen > self.FloaterTopHeight:
                    self.FloaterDiaClean = self.CheckModel[self.FloaterObj1].CylinderOuterDiameter[i]
                    break
    
            self.FloaterDiaMG = self.CheckModel[self.FloaterObj1].CylinderOuterDiameter[-2]
            self.FloaterMGThk = 1000*(self.FloaterDiaMG-self.FloaterDiaClean)/2

            self.T2_Line_Type  = self.CheckModel[self.CheckModel['Beam_' + self.minT2 + '_0_0'].LineType[0]]
            self.T1_Line_Type  = self.CheckModel[self.CheckModel['Beam_' + self.minT1 + '_0_0'].LineType[0]]
            # Temporary assuming CD as beam width
            print("The manually inputted beam width is 1.8m!")
            self.T2_Beam_width = 1.8 #self.T2_Line_Type.NormalDragLiftDiameter
            self.T1_Beam_width = 1.8 #self.T1_Line_Type.NormalDragLiftDiameter

            for i,obj in enumerate(self.Objects):
                if obj.typeName == 'Clump type' and obj.name.startswith('MooringBuoy'):    
                    self.BuoyHeight = self.CheckModel[obj.name].Height
                    self.NetActualBuoy = self.CheckModel[obj.name].Volume*1.025-self.CheckModel[obj.name].Mass
                    self.tmp = obj.name.split('_')
                    self.NetCleanBuoy = float(self.tmp[1][:-1])
                    break
                
            self.first_M1 = ''
            self.first_M2 = ''
            M1_Count = False
            M2_Count = False
            for i,obj in enumerate(self.Objects):
                if obj.typeName == 'Line' and obj.name.startswith('M1') and not M1_Count:
                    self.first_M1 = obj.name
                    M1_Count = True
                elif obj.typeName == 'Line' and obj.name.startswith('M2') and not M2_Count:
                    self.first_M2 = obj.name
                    M2_Count = True
                if M1_Count and M2_Count: break
                    
            self.MainLinePolyLength  = self.CheckModel[self.first_M1].Length[0]
            self.MainLineChainLength = self.CheckModel[self.first_M1].Length[1]
            self.MainLinePolyDia     = float(re.findall(r'\d+', self.CheckModel[self.first_M1].LineType[0])[0])
            self.MainLinePolyMG      = (1000*self.CheckModel[self.CheckModel[self.first_M1].LineType[0]].OD-self.MainLinePolyDia)/2
            self.MainLineChainDia    = float(re.findall(r'\d+', self.CheckModel[self.first_M1].LineType[1])[0])
            #
            if self.first_M2 != '':
                self.Top_Segment_Length = self.CheckModel[self.first_M2].Length[0] 
            else:
                self.Top_Segment_Length = self.CheckModel[self.first_M1].Length[0]  
 
            self.Num_Beam_Per_Platform = 0
            for i, obj in enumerate(self.CheckModel['G_Platform_' + str(self.minPlt)].GroupChildren(recurse=True)):
                if obj.name.startswith('Beam_'):
                    self.Num_Beam_Per_Platform += 1      
                    
            self.summary_Xls()
            self.summary_print()

    def findPlatformNum(self):
        # find T2 platform numbers
        self.T2PlatformNum = [0]
        for i,obj in enumerate(self.Objects):
            if obj.typeName == 'Link':
                self.tmp = int(re.findall(r'\d+', obj.name)[0])
                if self.tmp != self.T2PlatformNum[-1]:
                    self.T2PlatformNum.append(self.tmp)
        self.T2PlatformNum = sorted(self.T2PlatformNum[1:])
        # print('T2=',self.T2PlatformNum[:])
        # find all platforms number
        self.PlatformsNum = [0]
        for i,obj in enumerate(self.Objects):
            if obj.typeName == '6D buoy' and obj.name.startswith('Platform'):
                self.tmp = int(re.findall(r'\d+', obj.name)[0])
                if self.tmp != self.PlatformsNum[-1]:
                    self.PlatformsNum.append(self.tmp)
        self.PlatformsNum = sorted(self.PlatformsNum[1:])

        self.T1PlatformNum = [x for x in self.PlatformsNum if x not in self.T2PlatformNum]
        # print('T1=',self.T1PlatformNum[:])
        # Mooring attached platforms
        self.MoorPlatformNum = []
        self.MooringName     = []
        for i,obj in enumerate(self.Objects):
            if obj.typeName == 'Line' and (obj.name.startswith("M1") or obj.name.startswith("M2")):
                self.MooringName.append(obj.name)
                self.tmp = int(re.findall(r'\d+', obj.name)[1])
                if not self.tmp in self.MoorPlatformNum:
                    self.MoorPlatformNum.append(self.tmp)
        self.MooringName = sorted(self.MooringName, key=lambda x: (int(x.split('_')[1]), int((x.split('_')[0]).replace('M',''))))
        self.MoorPlatformNum = sorted(self.MoorPlatformNum)
        
        self.GroupNames = []
        for i,obj in enumerate(self.Objects):
            if obj.typeName == 'Browser group':
                self.GroupNames.append(obj.name)
        return
        
    def distance(self, objName1, objName2, obj1_Node = 0, obj2_Node = 0, static_stage = False):        
        if static_stage:
            point1 = np.array(self.static_global_coord(objName1)[obj1_Node])
            point2 = np.array(self.static_global_coord(objName2)[obj2_Node])
            
        else:    
            point1 = np.array(self.global_coord(objName1)[obj1_Node])
            point2 = np.array(self.global_coord(objName2)[obj2_Node])
            
        dist = np.linalg.norm(point2      - point1) 
        horz = np.linalg.norm(point2[:2]  - point1[:2]) 
        
        return point2[0]-point1[0], point2[1]-point1[1], point2[2]-point1[2], dist, horz
 
    def Mooring_distance(self):        
        if self.MooringName[0].split('_')[1] == self.MooringName[1].split('_')[1]:
            # dX, dY, dZ, dist, horz = self.distance(self.MooringName[0],self.MooringName[1],1,0)
            bridal_width           = self.distance(self.MooringName[1],self.MooringName[2],0,0)[3]
            M2_length              = self.distance(self.MooringName[1],self.MooringName[1],0,1)[3]
            bridal_angle           = math.degrees(np.arcsin(bridal_width/M2_length))
            M2_1                   = np.array(self.global_coord(self.MooringName[1])[0])
            M2_2                   = np.array(self.global_coord(self.MooringName[2])[0])
            bridal_avg_end1        = 0.5*(M2_1+M2_2)
            M1_end2                = np.array(self.global_coord(self.MooringName[0])[1])
            dX, dY, dZ             = bridal_avg_end1[0] - M1_end2[0], bridal_avg_end1[1] - M1_end2[1], bridal_avg_end1[2] - M1_end2[2]
            dist                   = np.linalg.norm(bridal_avg_end1     - M1_end2) 
            horz                   = np.linalg.norm(bridal_avg_end1[:2] - M1_end2[:2])           
        else:
            dX, dY, dZ, dist, horz = self.distance(self.MooringName[0],self.MooringName[0],1,0)
            bridal_angle           = math.inf
            
        return dX, dY, dZ, dist, horz, bridal_angle

    def min_horiz_dist_between_platfroms(self):
        min_dist   = 10000
        ref1_plat  = self.minT2
        ref1_beams = []
        ref1_links = []
        # find list of platform beams
        for i, obj in enumerate(self.CheckModel['G_Platform_'+ref1_plat].GroupChildren(recurse=True)):
            if obj.name.startswith('Beam_' + ref1_plat + '_'):
                ref1_beams.append(obj.name)
            elif obj.typeName=='Link':
                ref1_links.append(obj.name)
        # find min dist between beams of other platfroms with ref platfrom
        ref2_beams = []
        tmp1_plt = re.findall(r'\d+', self.CheckModel[ref1_links[0]].EndAConnection)[0]
        tmp2_plt = re.findall(r'\d+', self.CheckModel[ref1_links[0]].EndBConnection)[0]
        
        if ref1_plat == tmp1_plt:
            ref2_plat = tmp2_plt
        else:
            ref2_plat = tmp1_plt

        for i, obj in enumerate(self.CheckModel['G_Platform_'+ref2_plat].GroupChildren(recurse=True)):
            if obj.name.startswith('Beam_' + ref2_plat + '_'):
                ref2_beams.append(obj.name)
        
        for  i, ref2_beam in enumerate(ref2_beams):
            for j, ref1_beam in enumerate(ref1_beams):
                min_dist1 = self.distance(ref1_beam, ref2_beam,0,0)[4]
                min_dist2 = self.distance(ref1_beam, ref2_beam,0,1)[4]
                min_dist3 = self.distance(ref1_beam, ref2_beam,1,0)[4]
                min_dist4 = self.distance(ref1_beam, ref2_beam,1,1)[4]
                min_dist  = min(min_dist, min_dist1, min_dist2, min_dist3, min_dist4)
        
        return min_dist            

    def numbers(self,string,typeName):
        count = 0
        for i,obj in enumerate(self.Objects):
            if obj.name.startswith(string) and obj.typeName==typeName:
                count += 1
        return count
    
    def XDir_distance_envelope(self, string, int_pnt = 0):
        XDirMax = -100000
        XDirMin =  100000
        for i,obj in enumerate(self.Objects):
            if obj.name.startswith(string):
                tmp = self.global_coord(obj.name)[int_pnt][0]                    
                if tmp >= XDirMax:
                    XDirMax = tmp
                elif tmp <= XDirMin:
                    XDirMin = tmp
        return XDirMax-XDirMin
 
    def YDir_distance_envelope(self, string, int_pnt = 0):
        YDirMax = -100000
        YDirMin =  100000
        for i,obj in enumerate(self.Objects):
            if obj.name.startswith(string):
                tmp = self.global_coord(obj.name)[int_pnt][1]                    
                if tmp >= YDirMax:
                    YDirMax = tmp
                elif tmp <= YDirMin:
                    YDirMin = tmp
        return YDirMax-YDirMin       

    def weight(self,intPlatformsNum):
        self.PlatformMass = []
        for intPlatformNum in intPlatformsNum:
            tmpMass = 0
            for i,obj in enumerate(self.Objects):
                if re.findall(r'\d+', obj.name):
                    if intPlatformNum == int(re.findall(r'\d+', obj.name)[0]):
                        if obj.typeName == '6D buoy':
                            tmpMass += self.CheckModel[obj.name].Mass
                        if obj.typeName == 'Line' and obj.name.startswith('Beam'):
                            tmpType = self.CheckModel[obj.name].LineType[0] # for now assuming single section
                            tmpMass += self.CheckModel[tmpType].MassPerUnitLength * self.CheckModel[obj.name].Length[0]
            self.PlatformMass.append(tmpMass)
        # print(self.PlatformMass)
        return round(sum(self.PlatformMass),2), [round(num, 2) for num in self.PlatformMass]     

    def global_coord(self, objName):
        if not isinstance(self.CheckModel[objName].PolarZ, float):
            self.CheckModel[objName].PolarReferenceOrigin = ('Global origin','Global origin')
            self.CheckModel[objName].PolarReferenceAxes = ('Global axes','Global axes')
            GZA = self.CheckModel[objName].PolarZ[0]
            GRA = self.CheckModel[objName].PolarR[0]
            GThetaA = self.CheckModel[objName].PolarTheta[0]
            GXA = GRA*np.cos(np.deg2rad(GThetaA))
            GYA = GRA*np.sin(np.deg2rad(GThetaA))
            
            GZB = self.CheckModel[objName].PolarZ[1]
            GRB = self.CheckModel[objName].PolarR[1]
            GThetaB = self.CheckModel[objName].PolarTheta[1]
            GXB = GRB*np.cos(np.deg2rad(GThetaB))
            GYB = GRB*np.sin(np.deg2rad(GThetaB))
        else:
            self.CheckModel[objName].PolarReferenceOrigin = ('Global origin')
            self.CheckModel[objName].PolarReferenceAxes = ('Global axes')  
            GZA = self.CheckModel[objName].PolarZ
            GRA = self.CheckModel[objName].PolarR
            GThetaA = self.CheckModel[objName].PolarTheta
            GXA = GRA*np.cos(np.deg2rad(GThetaA))
            GYA = GRA*np.sin(np.deg2rad(GThetaA))
            GZB = None
            GXB = None
            GYB = None

        return [[GXA, GYA, GZA], [GXB, GYB, GZB]]

    def static_global_coord(self, objName, run_static = False):
        
        obj = self.CheckModel[objName]
        if run_static:
            self.CheckModel.CalculateStatics()
        
        if obj.typeName == 'Line':
            parameters = ['X', 'Y', 'Z']
            run_B      = True
            
        GXA = obj.StaticResult(parameters[0], ofx.oeEndA)
        GYA = obj.StaticResult(parameters[1], ofx.oeEndA)
        GZA = obj.StaticResult(parameters[2], ofx.oeEndA)

        if run_B:
            GXB = obj.StaticResult(parameters[0], ofx.oeEndB)
            GYB = obj.StaticResult(parameters[1], ofx.oeEndB)
            GZB = obj.StaticResult(parameters[2], ofx.oeEndB)
        else:
            GXB = None
            GYB = None
            GZB = None

        if run_static:
            self.CheckModel.Reset()
    
        return [[GXA, GYA, GZA], [GXB, GYB, GZB]]

    def CreateGroups(self, PltNumbers):
        
            self.CheckModel.CreateObject(ofx.otBrowserGroup, 'Mooring')
            for Mooring in self.MooringName:
                self.CheckModel[Mooring].groupParent = self.CheckModel['Mooring']
                
            for PltNum in PltNumbers:
                G_Name = 'G_Platform_' + str(PltNum)
                self.CheckModel.CreateObject(ofx.otBrowserGroup, G_Name)
                for obj in self.Objects:
                    if re.findall(r'\d+', obj.name) and (obj.typeName == 'Line' or obj.typeName == '6D buoy' or obj.typeName == '3D buoy' or obj.typeName == 'Link' or obj.typeName == 'Constraint'):
                        if not (obj.typeName == 'Line' and obj.name.startswith('M')): # to exclude mooring lines
                            if PltNum == int(re.findall(r'\d+', obj.name)[0]):
                                # print(obj.name,' to ',G_Name)
                                self.CheckModel[obj.name].groupParent = self.CheckModel[G_Name]
                                                        
    def CollapsedGroups(self, PltNumbers):
        yml_file = os.path.join(self.LocSims, self.FileName)
        with open(yml_file, 'r') as file:
            lines = file.readlines()
        with open(yml_file, 'w') as file:
            file.writelines(lines[:-2]) 
            file.write('\n    Collapsed:')
            file.write('\n      - Variable data')
            file.write('\n      - Wing types')
            file.write('\n      - Line types')
            file.write('\n      - Clump types')
            file.write('\n      - Mooring')

            for PltNum in PltNumbers:
                G_Name = 'G_Platform_' + str(PltNum) 
                file.write('\n      - '+G_Name)
            file.write('\n...')
    
    def BeamProperties(self, Plt_Num, Beam_Num):
        
        Current_Beam = 0
        for i, obj in enumerate(self.CheckModel['G_Platform_' + str(self.minPlt)].GroupChildren(recurse=True)):
            if obj.name.startswith('Beam_'):
                Current_Beam += 1
                if Current_Beam == Beam_Num:
                    Line_Type  = self.CheckModel[self.CheckModel[obj.name].LineType[0]]
                    break
                
        return_dict =  {'Type_Name': Line_Type.Name, 'OD': Line_Type.OD, 'ID': Line_Type.ID, 'CGx': Line_Type.CentreOfMassx, 'CGy': Line_Type.CentreOfMassy,
                        'MLen': Line_Type.MassPerUnitLength, 'EIx': Line_Type.EIx, 'EIy': Line_Type.EIy, 'EA': Line_Type.EA, 'Poisson Ratio': Line_Type.PoissonRatio,
                        'GJ': Line_Type.GJ, 'TT-Coup': Line_Type.TensionTorqueCoupling, 'Cdx': Line_Type.Cdx, 'Cdy': Line_Type.Cdy, 'Cdz': Line_Type.Cdz,
                        'CDN': Line_Type.NormalDragLiftDiameter, 'CDA': Line_Type.AxialDragLiftDiameter, 'Cax': Line_Type.Cax, 'Cay': Line_Type.Cay, 'Caz': Line_Type.Caz,
                        'Cmx': Line_Type.Cmx, 'Cmy': Line_Type.Cmy, 'Cmz': Line_Type.Cmz, 'Cs': Line_Type.Cs, 'Ce': Line_Type.Ce, 'Damp': Line_Type.RayleighDampingCoefficients}        
        
        return return_dict

    def GroupsProperties(self, PltNum):
        G_Name = 'G_Platform_' + str(PltNum)
        properties = ofx.CompoundProperties((self.CheckModel[G_Name],))
        #       Mass, CGX, CGY, CGZ, IXX, IYY, IZZ, Vol, VolX, VolY, VolZ
        return  (properties.Mass, properties.CentreOfMass[0], properties.CentreOfMass[1], properties.CentreOfMass[2],
                properties.MassMomentOfInertia[0,0], properties.MassMomentOfInertia[1,1], properties.MassMomentOfInertia[2,2],
                properties.Volume, properties.CentreOfVolume[0], properties.CentreOfVolume[1], properties.CentreOfVolume[2])
    
    def ConstraintHidden(self):
        for obj in self.Objects:
            if obj.typeName == 'Constraint':
                obj.Hidden = "True"    
        
    def summary_Xls(self):
        print('Calculating check excel summary data...')
        self.df1 = pd.DataFrame(self.PlatformsNum, columns=['Plt Num'])

        self.df1['Type1']                           = self.df1['Plt Num'].apply(lambda x: 'T1' if x in self.T1PlatformNum else 'T2')
        self.df1['Type2']                           = self.df1['Plt Num'].apply(lambda x: 'Mooring' if x in self.MoorPlatformNum else '-')
        self.df1['Weight']                          = self.weight(self.PlatformsNum)[1]
        self.df1['Weight2']                         = self.df1['Plt Num'].apply(lambda x: self.GroupsProperties(x)[0]).round(2)
        self.df1['CGX']                             = self.df1['Plt Num'].apply(lambda x: self.GroupsProperties(x)[1]).round(3)
        self.df1['CGY']                             = self.df1['Plt Num'].apply(lambda x: self.GroupsProperties(x)[2]).round(3)
        self.df1['CGZ']                             = self.df1['Plt Num'].apply(lambda x: self.GroupsProperties(x)[3]).round(3)
        self.df1['Ixx']                             = self.df1['Plt Num'].apply(lambda x: self.GroupsProperties(x)[4]).round(2)
        self.df1['Iyy']                             = self.df1['Plt Num'].apply(lambda x: self.GroupsProperties(x)[5]).round(2)
        self.df1['Izz']                             = self.df1['Plt Num'].apply(lambda x: self.GroupsProperties(x)[6]).round(2)
        self.df1['Volume']                          = self.df1['Plt Num'].apply(lambda x: self.GroupsProperties(x)[7]).round(2)
        self.df1['VolX']                            = self.df1['Plt Num'].apply(lambda x: self.GroupsProperties(x)[8]).round(3)
        self.df1['VolY']                            = self.df1['Plt Num'].apply(lambda x: self.GroupsProperties(x)[9]).round(3)
        self.df1['VolZ']                            = self.df1['Plt Num'].apply(lambda x: self.GroupsProperties(x)[10]).round(3)
        
        unit_df1 = [['[-]', '[-]', '[-]', '[MT]', '[MT]', '[m]', '[m]', '[m]', '[te.m^2]', '[te.m^2]', '[te.m^2]', '[m^3]', '[m]', '[m]', '[m]']]
        unit_df1 = pd.DataFrame(unit_df1, columns=self.df1.columns)
        self.df1 = pd.concat([unit_df1, self.df1]).reset_index(drop=True)
        self.TotalMoorPlatfroms = (self.df1['Type2'] == 'Mooring').sum()

        self.df2 = pd.DataFrame(self.PlatformsNum, columns=['Plt Num'])
        
        for iBeam in range(self.Num_Beam_Per_Platform):
            iBeam += 1
            self.df2['Line Type-' + str(iBeam)]                       = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Type_Name'])
            self.df2['OD-' + str(iBeam)]                              = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['OD']).round(3)
            self.df2['ID-' + str(iBeam)]                              = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['ID']).round(3)
            self.df2['CGX-' + str(iBeam)]                             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['CGx']).round(3)
            self.df2['CGY-' + str(iBeam)]                             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['CGy']).round(3)
            self.df2['Mass Per Unit Length-' + str(iBeam)]            = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['MLen']).round(3)
            self.df2['EIx-' + str(iBeam)]                             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['EIx']).round(3)
            self.df2['EIy-' + str(iBeam)]                             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['EIy']).round(3)
            self.df2['EA-' + str(iBeam)]                              = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['EA']).round(3)
            self.df2['Poisson Ratio-' + str(iBeam)]                   = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Poisson Ratio']).round(3)
            self.df2['GJ-' + str(iBeam)]                              = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['GJ']).round(3)
            self.df2['Tension-Torque Coup.-' + str(iBeam)]            = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['TT-Coup']).round(3)
            self.df2['Cdx-' + str(iBeam)]                             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Cdx']).round(3)
            self.df2['Cdy-' + str(iBeam)]                             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Cdy']).round(3)
            self.df2['Cdz-' + str(iBeam)]                             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Cdz']).round(3)
            self.df2['Normal Drag-Lift Dia-' + str(iBeam)]            = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['CDN']).round(3)
            self.df2['Axial Drag-Lift Dia-' + str(iBeam)]             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['CDA']).round(3)
            self.df2['Cax-' + str(iBeam)]                             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Cax']).round(3)
            self.df2['Cay-' + str(iBeam)]                             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Cay']).round(3)
            self.df2['Caz-' + str(iBeam)]                             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Caz']).round(3)
            self.df2['Cmx-' + str(iBeam)]                             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Cmx']).round(3)
            self.df2['Cmy-' + str(iBeam)]                             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Cmy']).round(3)
            self.df2['Cmz-' + str(iBeam)]                             = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Cmz']).round(3)
            self.df2['Cs-' + str(iBeam)]                              = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Cs']).round(3)
            self.df2['Ce-' + str(iBeam)]                              = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Ce']).round(3)
            self.df2['Damping Coef-' + str(iBeam)]                    = self.df2['Plt Num'].apply(lambda x: self.BeamProperties(x, iBeam)['Damp'])

        unit_df2 = [['[-]', '[m]', '[m]', '[m]', '[m]', '[te/m]', '[kN.m^2]', '[kN.m^2]', '[kN]',
                     '[-]', '[kN.m^2]', '[kN.m]', '[-]', '[-]', '[-]', '[m]', '[m]', '[-]', '[-]',
                     '[-]', '[-]', '[-]', '[-]', '[-]', '[-]', '[-]']*self.Num_Beam_Per_Platform]
        
        unit_df2 = [['[-]'] + inner for inner in unit_df2]
        unit_df2 = pd.DataFrame(unit_df2, columns=self.df2.columns)
        self.df2 = pd.concat([unit_df2, self.df2]).reset_index(drop=True)

        data = [
            {'Item': 'Water Depth',                         'Value': self.WaterDepth,                                                                               'Unit': '[m]'},
            {'Item': 'Number of Platforms',                 'Value': self.numbers('Floater', '6D buoy')/3,                                                          'Unit': '[-]'},
            {'Item': 'Total Platforms Mass',                'Value': self.weight(self.PlatformsNum)[0],                                                             'Unit': '[MT]'},
            {'Item': 'Number of T1 Platforms',              'Value': len(self.T1PlatformNum),                                                                       'Unit': '[-]'},
            {'Item': 'Number of T2 Platforms',              'Value': len(self.T2PlatformNum),                                                                       'Unit': '[-]'},
            {'Item': 'Num. of Mooring Attached Plt.',       'Value': self.TotalMoorPlatfroms,                                                                       'Unit': '[-]'},
            {'Item': '1st T1 Platform Mass',                'Value': self.weight([min(self.T1PlatformNum)])[0],                                                     'Unit': '[MT]'},
            {'Item': '1st T2 Platform Mass',                'Value': self.weight([min(self.T2PlatformNum)])[0],                                                     'Unit': '[MT]'},
            {'Item': 'X-Dir Platforms Width (c-c Float.)',  'Value': self.XDir_distance_envelope('Floater'),                                                        'Unit': '[m]'},
            {'Item': 'Y-Dir Platforms Width (c-c Float.)',  'Value': self.YDir_distance_envelope('Floater'),                                                        'Unit': '[m]'},
            {'Item': '1st T1 Platform Truss Width',         'Value': self.T1_Beam_width,                                                                            'Unit': '[m]'},
            {'Item': '1st T2 Platform Truss Width',         'Value': self.T2_Beam_width,                                                                            'Unit': '[m]'},            
            {'Item': 'Gap between Platforms Beams C-C',     'Value': self.min_horiz_dist_between_platfroms(),                                                       'Unit': '[m]'},
            {'Item': 'Gap between Platforms Beams S-S',     'Value': self.min_horiz_dist_between_platfroms()-self.T2_Beam_width/2-self.T1_Beam_width/2,             'Unit': '[m]'},
            {'Item': 'Floaters C-C Distance',               'Value': self.distance('Floater'+self.minT2+'_0','Floater'+self.minT2+'_120')[3],                       'Unit': '[m]'},
            {'Item': 'Floater Bottom Height',               'Value': self.FloaterBottomHeight,                                                                      'Unit': '[m]'},
            {'Item': 'Floater Clean Diameter',              'Value': self.FloaterDiaClean,                                                                          'Unit': '[m]'},
            {'Item': 'Floater MG Thickness',                'Value': self.FloaterMGThk,                                                                             'Unit': '[mm]'},
            {'Item': 'Truss axis height from top platform', 'Value': self.distance('Platform_'+self.minT2+'_0','Beam_'+self.minT2+'_0_0',0,1)[2],                   'Unit': '[m]'},
            {'Item': 'Platform Fairleads horz. dist',       'Value': self.distance('Con_Frld_'+self.minMoor+'_240_1','Con_Frld_'+self.minMoor+'_240_-1',0,0)[4],    'Unit': '[m]'},
            {'Item': 'X-Dir Mooring Foortprint Width',      'Value': self.XDir_distance_envelope('M1', 1),                                                          'Unit': '[m]'},
            {'Item': 'Y-Dir Mooring Foortprint Width',      'Value': self.YDir_distance_envelope('M1', 1),                                                          'Unit': '[m]'},
            {'Item': '1st Mooring Leg Horizontal Dist.',    'Value': self.Mooring_distance()[4],                                                                    'Unit': '[m]'},
            {'Item': '1st Mooring Leg Distance',            'Value': self.Mooring_distance()[3],                                                                    'Unit': '[m]'},
            {'Item': '1st Mooring Bridal Angle (initial)',  'Value': self.Mooring_distance()[5],                                                                    'Unit': '[deg]'},
            {'Item': 'Number of Main Lines (M1)',           'Value': self.numbers('M1','Line'),                                                                     'Unit': '[-]'},
            {'Item': 'Number of Secondary Lines (M2)',      'Value': self.numbers('M2','Line'),                                                                     'Unit': '[-]'},
            {'Item': 'Top Line Segment Length',             'Value': self.Top_Segment_Length,                                                                       'Unit': '[m]'},
            {'Item': 'Main Mooring Polyester Dia',          'Value': self.MainLinePolyDia,                                                                          'Unit': '[mm]'},
            {'Item': 'Main Mooring Polyester MG Thk',       'Value': self.MainLinePolyMG,                                                                           'Unit': '[mm]'},
            {'Item': 'Main Mooring Polyester Length',       'Value': self.MainLinePolyLength,                                                                       'Unit': '[m]'},
            {'Item': 'Main Mooring Chain Clean Dia',        'Value': self.MainLineChainDia,                                                                         'Unit': '[mm]'},
            {'Item': 'Main Mooring Chain Length',           'Value': self.MainLineChainLength,                                                                      'Unit': '[m]'},
            {'Item': 'Buoy Clean Net Buoyancy',             'Value': self.NetCleanBuoy,                                                                             'Unit': '[MT]'},
            {'Item': 'Buoy Actual Net Buoyancy',            'Value': self.NetActualBuoy,                                                                            'Unit': '[MT]'},
            {'Item': 'Buoy Height',                         'Value': self.BuoyHeight,                                                                               'Unit': '[m]'}
        ]

        self.df3 = pd.DataFrame(data)
        self.df3['Value'] = self.df3['Value'].round(2)

        data = [
            {'Item': 'In-Out Coupling ends A horz. dist',     'Value': self.distance('In_'+self.minT2+'_0_-1','Out_'+self.minT2+'_0_-1',0,0)[4],                        'Unit': '[m]', 'IV Description': '3200 Coupling Width - View E'},
            {'Item': 'In-Out Coupling ends A vert. dist',     'Value': self.distance('In_'+self.minT2+'_0_-1','Out_'+self.minT2+'_0_-1',0,0)[2],                        'Unit': '[m]', 'IV Description': ''},
            {'Item': 'In Cylinder horz. len.',                'Value': self.distance('In_'+self.minT2+'_0_-1','In_'+self.minT2+'_0_-1',0,1)[4],                         'Unit': '[m]', 'IV Description': ''},
            {'Item': 'In Cylinder length',                    'Value': self.distance('In_'+self.minT2+'_0_-1','In_'+self.minT2+'_0_-1',0,1)[3],                         'Unit': '[m]', 'IV Description': 'Cylinder length 4282 - View E'},
            {'Item': 'In Cyl. endA X dist to truss side',     'Value': self.distance('Beam_'+self.minT2+'_0_0','In_'+self.minT2+'_0_1',1,0)[0]-self.T2_Beam_width/2,    'Unit': '[m]', 'IV Description': 'Coupling Pos X 300'},
            {'Item': 'In Cyl. endA Y dist to floater cen.',   'Value': self.distance('Floater'+self.minT2+'_0','In_'+self.minT2+'_0_1',0,0)[1],                         'Unit': '[m]', 'IV Description': ''},
            {'Item': 'In Cyl. endA Z dist to floater cen.',   'Value': self.distance('Floater'+self.minT2+'_0','In_'+self.minT2+'_0_1',0,0)[2],                         'Unit': '[m]', 'IV Description': ''},
            {'Item': 'In Cyl. endB Z vert to floater cen.',   'Value': self.distance('Floater'+self.minT2+'_0','In_'+self.minT2+'_0_1',0,1)[2],                         'Unit': '[m]', 'IV Description': 'Coupling Center Pos Z 2650' },
            {'Item': 'Out Cylinder horz. len.',               'Value': self.distance('Out_'+self.minT2+'_0_-1','Out_'+self.minT2+'_0_-1',0,1)[4],                       'Unit': '[m]', 'IV Description': ''},
            {'Item': 'Out Cylinder length',                   'Value': self.distance('Out_'+self.minT2+'_0_-1','Out_'+self.minT2+'_0_-1',0,1)[3],                       'Unit': '[m]', 'IV Description': 'Cylinder length 4282 - View E'},
            {'Item': 'Out Cyl. endA X dist to truss side',    'Value': self.distance('Beam_'+self.minT2+'_0_0','Out_'+self.minT2+'_0_1',1,0)[0]-self.T2_Beam_width/2,   'Unit': '[m]', 'IV Description': 'Coupling Pos X 300'},
            {'Item': 'Out Cyl. endA Y dist to floater cen.',  'Value': self.distance('Floater'+self.minT2+'_0','Out_'+self.minT2+'_0_1',0,0)[1],                        'Unit': '[m]', 'IV Description': ''},
            {'Item': 'Out Cyl endA Z dist to floater cen.',   'Value': self.distance('Floater'+self.minT2+'_0','Out_'+self.minT2+'_0_1',0,0)[2],                        'Unit': '[m]', 'IV Description': ''},
            {'Item': 'Out Cyl. endB Z vert to floater cen.',  'Value': self.distance('Floater'+self.minT2+'_0','Out_'+self.minT2+'_0_1',0,1)[2],                        'Unit': '[m]', 'IV Description': 'Coupling Center Pos Z 2650' },
            {'Item': 'In-Mid Coup ends A trans. horz. dist',  'Value': self.distance('In_'+self.minT2+'_0_-1','Mid_'+self.minT2+'_0_-1',0,0)[4],                        'Unit': '[m]', 'IV Description': ''},
            {'Item': 'In-Mid Coup ends A vert. dist',         'Value': abs(self.distance('In_'+self.minT2+'_0_-1','Mid_'+self.minT2+'_0_-1',0,0)[2]),                   'Unit': '[m]', 'IV Description': '3400 Coupling Pos Z'},
            {'Item': 'Out-Mid Coup ends A trans. horz. dist', 'Value': self.distance('Out_'+self.minT2+'_0_-1','Mid_'+self.minT2+'_0_-1',0,0)[4],                       'Unit': '[m]', 'IV Description': ''},
            {'Item': 'Out-Mid Coup ends A vert. dist',        'Value': abs(self.distance('Out_'+self.minT2+'_0_-1','Mid_'+self.minT2+'_0_-1',0,0)[2]),                  'Unit': '[m]', 'IV Description': '3400 Coupling Pos Z'},
            {'Item': 'Mid Link horz. len.',                   'Value': self.distance('Mid_'+self.minT2+'_0_-1','Mid_'+self.minT2+'_0_-1',0,1)[4],                       'Unit': '[m]', 'IV Description': ''},
            {'Item': 'Mid Link length',                       'Value': self.distance('Mid_'+self.minT2+'_0_-1','Mid_'+self.minT2+'_0_-1',0,1)[3],                       'Unit': '[m]', 'IV Description': '4667 Coupling Mid length'},
            {'Item': 'Mid Link endA X dist to truss side',    'Value': self.distance('Beam_'+self.minT2+'_0_0','Mid_'+self.minT2+'_0_1',1,0)[0]-self.T2_Beam_width/2,   'Unit': '[m]', 'IV Description': 'Coupling Mid X 379'},
            {'Item': 'Mid Link endA Y dist to floater cen.',  'Value': self.distance('Floater'+self.minT2+'_0','Mid_'+self.minT2+'_0_1',0,0)[1],                        'Unit': '[m]', 'IV Description': ''},
            {'Item': 'Mid Link endA Z dist to floater cen.',  'Value': self.distance('Floater'+self.minT2+'_0','Mid_'+self.minT2+'_0_1',0,0)[2],                        'Unit': '[m]', 'IV Description': 'Coupling Mid Z 30'},
            {'Item': 'Mid Link endB Z vert to floater cen.',  'Value': self.distance('Floater'+self.minT2+'_0','Mid_'+self.minT2+'_0_1',0,1)[2],                        'Unit': '[m]', 'IV Description': 'Coupling Center Pos Z 2650'},
            {'Item': 'Mid Links endA half dist',              'Value': abs(0.5*self.distance('Mid_'+self.minT2+'_0_-1','Mid_'+self.minT2+'_0_1',0,0)[3]),               'Unit': '[m]', 'IV Description': 'Coupling Center Pos Y 10900'},
        ]

        self.df4 = pd.DataFrame(data)
        self.df4['Value'] = self.df4['Value'].round(2)
        
        self.df5 = self.calc_pre_tension() # pre-tension sheet
        
        print('Writing check excel summary file...')
        if os.path.isfile(self.ExcelFile):
            with pd.ExcelWriter(self.ExcelFile, engine='openpyxl', mode='a') as writer:
                self.df1.to_excel(writer, sheet_name=self.SheetName+'-Wgt',     index=True)
                self.df2.to_excel(writer, sheet_name=self.SheetName+'-Beam',    index=True)
                self.df3.to_excel(writer, sheet_name=self.SheetName+'-General', index=True)
                self.df4.to_excel(writer, sheet_name=self.SheetName+'-Coupl',   index=True)
                self.df5.to_excel(writer, sheet_name=self.SheetName+'-pre_T',   index=True)
                self.Excel_Alignment(writer.book, CheckSummary.Report_Count)
        else:
            with pd.ExcelWriter(self.ExcelFile, engine='openpyxl', mode='w') as writer:
                self.df1.to_excel(writer, sheet_name=self.SheetName+'-Wgt',     index=True)
                self.df2.to_excel(writer, sheet_name=self.SheetName+'-Beam',    index=True)
                self.df3.to_excel(writer, sheet_name=self.SheetName+'-General', index=True)
                self.df4.to_excel(writer, sheet_name=self.SheetName+'-Coupl',   index=True)
                self.df5.to_excel(writer, sheet_name=self.SheetName+'-pre_T',   index=True)
                self.Excel_Alignment(writer.book, CheckSummary.Report_Count)
        
        return self.TotalMoorPlatfroms

    def calc_pre_tension(self):
        df5 = pd.DataFrame(columns=['Mooring Name','Length','Pre_Tension',
                                    'Initial Horiz. Dist.','Static Horiz. Dist.','Initial Distance','Static Distance',
                                    'End A GX','End A GY','End A GZ', 
                                    'End B GX','End B GY','End B GZ'])
        for iName, objName in enumerate(self.MooringName):
            Mooring_Object = self.CheckModel[objName]
            df5.loc[iName,'Mooring Name']                     = objName
            df5.loc[iName,'Length']                           = round(np.sum(Mooring_Object.Length), 2)
            df5.loc[iName,'Initial Horiz. Dist.']             = round(self.distance(objName,objName,0,1)[4], 2)
            df5.loc[iName,'Initial Distance']                 = round(self.distance(objName,objName,0,1)[3], 2)
            df5.loc[iName,['End A GX','End A GY','End A GZ']] = [round(x, 2) for x in self.global_coord(objName)[0]]
            df5.loc[iName,['End B GX','End B GY','End B GZ']] = [round(x, 2) for x in self.global_coord(objName)[1]]
            
        self.CheckModel.CalculateStatics()
        for iName, objName in enumerate(self.MooringName): # because after static calc can't load Polar coordinates!
            Mooring_Object = self.CheckModel[objName]
            df5.loc[iName,'Pre_Tension']                      = round(Mooring_Object.StaticResult('Effective tension', ofx.oeEndA), 2)
            df5.loc[iName,'Static Horiz. Dist.']              = round(self.distance(objName,objName,0,1, static_stage = True)[4], 2)
            df5.loc[iName,'Static Distance']                  = round(self.distance(objName,objName,0,1, static_stage = True)[3], 2)

        unit_df5 = [['[-]', '[m]', '[kN]', '[m]', '[m]', '[m]', '[m]', '[m]', '[m]', '[m]', '[m]', '[m]', '[m]']]
        unit_df5 = pd.DataFrame(unit_df5, columns=df5.columns)
        df5 = pd.concat([unit_df5, df5]).reset_index(drop=True)            
            
        self.CheckModel.Reset()
        
        return df5
    
    def Excel_Alignment(self, workbook, Report_Count):

        for sheet_name in workbook.sheetnames:
            if int(sheet_name.split('-')[0]) != Report_Count: continue
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        return None
        
    def summary_print(self):
        print(f"\n {self.SheetName} Model General Checking Summary:")
        print(self.df3)
        print(f"\n {self.SheetName} Coupling Checking Summary:")
        print(self.df4.iloc[:, 0:3])
        print("\n")

    def Anchor_Allowables(self, allowableGX = 1000, allowableGZ = 150):
        
        self.CheckModel['General'].NumberOfUserDefinedResults = 1
        self.CheckModel['General'].UserDefinedResultName[0] = 'allowable GX'
        self.CheckModel['General'].UserDefinedResultScriptFileSource = 'Embedded' # External file
        
        scripts__allowable_GX = [
        
        "import numpy \n",
                                 
        "def endGXallowable(info, allowableGX = ", str(allowableGX), """):
            if info.modelObject.TopEnd == "End A":
                endLoadGX = info.modelObject.TimeHistory("End GX force", info.period, OrcFxAPI.oeEndB)
            elif info.modelObject.TopEnd == "End B":
                endLoadGX = info.modelObject.TimeHistory("End GX force", info.period, OrcFxAPI.oeEndA)
            return numpy.abs(endLoadGX)/allowableGX \n""",
        
        """def UserDefinedResults(model):
            return (
                {
                    "ObjectType": OrcFxAPI.ObjectType.Line,
                    "Name": "Anchor End GX allowable",
                    "Units": "FF/FF",
                    "LineResultPoints": OrcFxAPI.LineResultPoints.WholeLine,
                    "TimeDomainFunction": endGXallowable
                },
            )"""                    
        ]
            
        self.CheckModel['General'].UserDefinedResultScript = ''.join(scripts__allowable_GX)

        self.CheckModel['General'].NumberOfUserDefinedResults = 2
        self.CheckModel['General'].UserDefinedResultName[1] = 'allowable GZ'
        self.CheckModel['General'].UserDefinedResultScriptFileSource = 'Embedded' # External file
        
        scripts__allowable_GZ = [
        
        "import numpy \n",
                                 
        "def endGZallowable(info, allowableGZ = ", str(allowableGZ), """):
            if info.modelObject.TopEnd == "End A":
                endLoadGZ = info.modelObject.TimeHistory("End GZ force", info.period, OrcFxAPI.oeEndB)
            elif info.modelObject.TopEnd == "End B":
                endLoadGZ = info.modelObject.TimeHistory("End GZ force", info.period, OrcFxAPI.oeEndA)
            
            endLoadGZ = numpy.array([0 if value > 0 else -value for value in endLoadGZ])
            return endLoadGZ/allowableGZ \n""",
        
        """def UserDefinedResults(model):
            return (
                {
                    "ObjectType": OrcFxAPI.ObjectType.Line,
                    "Name": "Anchor End GZ allowable",
                    "Units": "FF/FF",
                    "LineResultPoints": OrcFxAPI.LineResultPoints.WholeLine,
                    "TimeDomainFunction": endGZallowable
                },
            )"""                    
        ]
    
        self.CheckModel['General'].UserDefinedResultScript = ''.join(scripts__allowable_GZ)
                 
class General: 
    def __init__(self):
        self.StaticsMaxDamping=80
        self.StaticsMinDamping=10
        self.StaticsMaxIterations=2000
        self.ImplicitConstantTimeStep=0.05
        self.DurationSimulation=3600
        self.TargetLogSampleInterval=0.1


